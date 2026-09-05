# v262
#!/usr/bin/env python3
"""vys-262 Render #2 heavy worker · Пер-R27.

Responsibilities:
- mutual peer health ping with Render #1;
- receive compact SQLite page deltas from front and validate reconstructed state;
- keep a Redis delta journal plus local exact restore cache;
- generate periodic full Redis/MEGA checkpoints locally;
- serve cached latest snapshot to Render #1 for fast deploy restore;
- execute Google Sheets creation (OAuth + Sheets API) away from Telegram frontend.
"""
from __future__ import annotations
import base64
import csv
import mimetypes
import gzip
import hashlib
import json
import os
import queue
import re
import secrets
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
import requests
try:
    import redis as _redis
except Exception:
    _redis = None
from flask import Flask, request, Response, send_file
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from runtime_config import install_internal_runtime_config, CONFIG_VERSION as INTERNAL_CONFIG_VERSION
install_internal_runtime_config("worker")

app = Flask(__name__)
VERSION = 'vys-262-worker-per-r27-heavy'
TRANSPORT_VERSION = 'vys-262-worker-per-r27-internal-config'


def env_bool(name, default=False):
    return str(os.getenv(name, '1' if default else '0') or '').strip().lower() in {'1','true','yes','on','да'}

def env_int(name, default, lo, hi):
    try: return max(lo, min(hi, int(os.getenv(name, str(default)) or str(default))))
    except Exception: return default

def peer_secret(): return str(os.getenv('PEER_SHARED_SECRET','') or '').strip()
def authorized():
    expected, supplied = peer_secret(), str(request.headers.get('X-Peer-Secret','') or '')
    return bool(expected and secrets.compare_digest(expected, supplied))
def front_base():
    raw = str(os.getenv('FRONT_SERVICE_URL', os.getenv('PEER_SERVICE_URL','')) or '').strip().rstrip('/')
    if raw and not raw.startswith(('http://','https://')): raw = 'https://' + raw
    return raw
def mega_root(): return '/' + str(os.getenv('MEGA_BACKUP_DIR','TelegramBotBackups2-2') or 'TelegramBotBackups2-2').strip('/')
def remote_db_dir(): return mega_root().rstrip('/') + '/database'
def remote_latest(): return remote_db_dir().rstrip('/') + '/latest_bot_state.sqlite3.gz'
def remote_history_dir(): return remote_db_dir().rstrip('/') + '/history'

STATE_LOCK = threading.RLock()
MEGA_LOCK = threading.RLock()
GOOGLE_LOCK = threading.RLock()
STATE = {
    'started_at': time.time(), 'peer_last_attempt':0.0, 'peer_last_ok':0.0, 'peer_last_error':'', 'peer_status':None,
    'front_health':{},
    'job_last_id':'', 'job_last_type':'', 'job_last_reason':'', 'job_last_started':0.0, 'job_last_done':0.0, 'job_last_error':'',
    'sync_count':0, 'sync_failures':0, 'last_snapshot_sha256':'', 'last_snapshot_size':0, 'last_snapshot_at':0.0, 'last_full_fetch_started':0.0, 'full_fetch_suppressed':0,
    'last_mega_upload_at':0.0, 'last_restore_download_at':0.0, 'mega_layout_ok':False, 'mega_warmup_ok':False,
    'google_jobs':0, 'google_failures':0, 'google_last_ok':0.0, 'google_last_error':'',
    'cache_revision':0.0,
    'last_state_token':'', 'active_state_token':'', 'dirty_state_token':'', 'deduped_sync_requests':0,
    'redis_cache_ok':False, 'redis_last_write':0.0, 'redis_last_read':0.0, 'redis_last_error':'',
    'delta_count':0, 'delta_since_checkpoint':0, 'delta_bytes':0, 'delta_last_at':0.0, 'delta_last_pages':0, 'delta_last_error':'',
    'delta_replayed':0, 'full_checkpoint_at':0.0, 'full_checkpoint_count':0, 'last_state_sha256':'',
    'delta_bytes_since_checkpoint':0,
    'event_received':0, 'event_committed':0, 'event_mirrored':0, 'event_pending':0, 'event_last_at':0.0, 'event_last_error':'',
    'reconcile_last_at':0.0, 'reconcile_last_ok':0.0, 'reconcile_last_error':'', 'reconcile_full_resyncs':0,
}
JOB_Q = queue.Queue(maxsize=32)
GOOGLE_Q = queue.Queue(maxsize=16)
GOOGLE_JOB_LOCK = threading.RLock()
GOOGLE_JOB_STATUS = {}
SYNC_PENDING_LOCK = threading.RLock(); SYNC_PENDING = False; SYNC_DIRTY = False; SYNC_DIRTY_REASON = ''
RESTORE_REFRESH_LOCK = threading.RLock(); RESTORE_REFRESH_RUNNING = False
CACHE_DIR = Path(os.getenv('WORKER_CACHE_DIR','/tmp/vys262_worker') or '/tmp/vys262_worker'); CACHE_DIR.mkdir(parents=True, exist_ok=True)
CACHE_LATEST = CACHE_DIR / 'latest_bot_state.sqlite3.gz'
CACHE_DB = CACHE_DIR / 'latest_bot_state.sqlite3'
_GOOGLE_TOKEN = {'token':'','expires_at':0.0}
_REDIS_CLIENT = None
_REDIS_LOCK = threading.RLock()
_REDIS_SNAPSHOT_KEY = str(os.getenv('WORKER_REDIS_SNAPSHOT_KEY','vys262:bot_state:latest_gz') or 'vys262:bot_state:latest_gz').strip()
_REDIS_META_KEY = _REDIS_SNAPSHOT_KEY + ':meta'
_REDIS_DELTA_KEY = str(os.getenv('WORKER_REDIS_DELTA_KEY', _REDIS_SNAPSHOT_KEY + ':deltas_v1') or (_REDIS_SNAPSHOT_KEY + ':deltas_v1')).strip()
_REDIS_DELTA_META_KEY = _REDIS_DELTA_KEY + ':meta'
_REDIS_EVENT_PREFIX = str(os.getenv('WORKER_REDIS_EVENT_PREFIX','vys262:tg_events:v1') or 'vys262:tg_events:v1').strip()
_REDIS_CAPSULE_KEY = str(os.getenv('WORKER_REDIS_CAPSULE_KEY','vys262:durable_capsule:r20') or 'vys262:durable_capsule:r20').strip()
CAPSULE_LOCAL = CACHE_DIR / 'durable_capsule_r20.json.gz'
CAPSULE_LOCK = threading.RLock()
CAPSULE_MEGA_Q = queue.Queue(maxsize=1)
CAPSULE_MEGA_STATE = {'last_ok':0.0,'last_error':'','uploads':0,'restores':0}
EVENT_DB = CACHE_DIR / 'event_journal.sqlite3'
EVENT_LOCK = threading.RLock()
EVENT_REDIS_Q = queue.Queue(maxsize=env_int('WORKER_EVENT_REDIS_QUEUE_MAX',2048,64,20000))

def _event_db_init_v268():
    with EVENT_LOCK:
        conn=sqlite3.connect(EVENT_DB,timeout=10)
        try:
            conn.execute('PRAGMA journal_mode=WAL')
            conn.execute('PRAGMA synchronous=FULL')
            conn.execute("CREATE TABLE IF NOT EXISTS events(event_id TEXT PRIMARY KEY, update_id TEXT, chat_id TEXT, update_type TEXT, payload_json TEXT, payload_sha256 TEXT, state TEXT, received_at REAL, committed_at REAL, mirrored_at REAL, state_token TEXT, last_error TEXT, updated_at REAL)")
            conn.execute('CREATE INDEX IF NOT EXISTS idx_events_state_time ON events(state,updated_at)')
            conn.commit()
        finally: conn.close()

def _event_local_upsert_v268(row:dict):
    if not isinstance(row,dict): return False
    eid=str(row.get('event_id') or row.get('update_id') or '')
    if not eid: return False
    _event_db_init_v268()
    with EVENT_LOCK:
        conn=sqlite3.connect(EVENT_DB,timeout=10)
        try:
            old=conn.execute('SELECT payload_json,payload_sha256,state,received_at,committed_at,mirrored_at,state_token,last_error FROM events WHERE event_id=?',(eid,)).fetchone()
            payload=row.get('payload')
            payload_json=json.dumps(payload,ensure_ascii=False,separators=(',',':'),default=str) if isinstance(payload,dict) else (old[0] if old else '{}')
            incoming_state=str(row.get('state') or (old[2] if old else 'received'))
            old_state=str(old[2] if old else '')
            _rank={'received':1,'failed_retry':1,'committed':2,'mirrored':3,'checkpointed':4,'done':4}
            state=old_state if _rank.get(old_state,0)>_rank.get(incoming_state,0) else incoming_state
            received=float(row.get('received_at') or (old[3] if old else time.time()) or time.time())
            committed=float(row.get('committed_at') or (old[4] if old else 0.0) or 0.0)
            mirrored=float(row.get('mirrored_at') or (old[5] if old else 0.0) or 0.0)
            token=str(row.get('state_token') or (old[6] if old else '') or '')[:120]
            err=str(row.get('last_error') or (old[7] if old else '') or '')[:300]
            conn.execute("INSERT INTO events(event_id,update_id,chat_id,update_type,payload_json,payload_sha256,state,received_at,committed_at,mirrored_at,state_token,last_error,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(event_id) DO UPDATE SET update_id=excluded.update_id,chat_id=excluded.chat_id,update_type=excluded.update_type,payload_json=CASE WHEN excluded.payload_json!='{}' THEN excluded.payload_json ELSE events.payload_json END,payload_sha256=CASE WHEN excluded.payload_sha256!='' THEN excluded.payload_sha256 ELSE events.payload_sha256 END,state=excluded.state,received_at=events.received_at,committed_at=CASE WHEN excluded.committed_at>0 THEN excluded.committed_at ELSE events.committed_at END,mirrored_at=CASE WHEN excluded.mirrored_at>0 THEN excluded.mirrored_at ELSE events.mirrored_at END,state_token=CASE WHEN excluded.state_token!='' THEN excluded.state_token ELSE events.state_token END,last_error=excluded.last_error,updated_at=excluded.updated_at",
                (eid,str(row.get('update_id') or eid),str(row.get('chat_id') if row.get('chat_id') is not None else ''),str(row.get('update_type') or 'other')[:40],payload_json,str(row.get('payload_sha256') or (old[1] if old else '') or ''),state,received,committed,mirrored,token,err,time.time()))
            conn.commit(); return True
        finally: conn.close()

def _event_redis_store_v268(row:dict):
    client=_redis_client()
    if client is None: return False,'REDIS_URL not configured'
    eid=str(row.get('event_id') or row.get('update_id') or '')
    if not eid: return False,'event id empty'
    try:
        key=f'{_REDIS_EVENT_PREFIX}:event:{eid}'; pending=f'{_REDIS_EVENT_PREFIX}:pending'
        current={}
        raw=client.get(key)
        if raw:
            try: current=json.loads(raw.decode('utf-8') if isinstance(raw,(bytes,bytearray)) else raw)
            except Exception: current={}
        merged=dict(current or {}); merged.update(row)
        _rank={'received':1,'failed_retry':1,'committed':2,'mirrored':3,'checkpointed':4,'done':4}
        old_state=str((current or {}).get('state') or ''); new_state=str((row or {}).get('state') or '')
        if _rank.get(old_state,0)>_rank.get(new_state,0): merged['state']=old_state
        merged['updated_at']=time.time()
        ttl=env_int('WORKER_EVENT_RETENTION_SEC',604800,86400,2592000)
        pipe=client.pipeline(transaction=True)
        pipe.set(key,json.dumps(merged,ensure_ascii=False,separators=(',',':'),default=str),ex=ttl)
        if str(merged.get('state') or '') in {'mirrored','checkpointed','done'}: pipe.zrem(pending,eid)
        else: pipe.zadd(pending,{eid:float(merged.get('received_at') or time.time())})
        pipe.execute(); return True,'Redis event stored'
    except Exception as exc: return False,f'{type(exc).__name__}: {str(exc)[:220]}'



def _event_redis_enqueue_v270(row:dict) -> bool:
    """R15: acknowledge the remote witness after Worker-local fsync; Redis flush is detached.

    This keeps Render #1 fast while retaining a second-service copy immediately.  The
    reconcile loop retries every unmirrored local event into Redis until it succeeds.
    """
    try:
        EVENT_REDIS_Q.put_nowait(dict(row or {}))
        return True
    except queue.Full:
        return False

def _event_redis_flush_loop_v270():
    while True:
        row = EVENT_REDIS_Q.get()
        try:
            ok, detail = _event_redis_store_v268(row)
            if not ok:
                with STATE_LOCK:
                    STATE['event_last_error'] = ('async redis: ' + str(detail))[:220]
                # Small retry without blocking the HTTP receipt path.
                time.sleep(env_int('WORKER_EVENT_REDIS_RETRY_MS',250,10,5000)/1000.0)
                try: EVENT_REDIS_Q.put_nowait(row)
                except queue.Full: pass
        except Exception as exc:
            with STATE_LOCK:
                STATE['event_last_error'] = f'async redis {type(exc).__name__}: {str(exc)[:180]}'
        finally:
            EVENT_REDIS_Q.task_done()

def _event_hydrate_pending_from_redis_v268(limit=250):
    client=_redis_client()
    if client is None: return 0
    loaded=0
    try:
        ids=client.zrange(f'{_REDIS_EVENT_PREFIX}:pending',0,max(0,min(999,int(limit)-1))) or []
        for raw_id in ids:
            eid=raw_id.decode() if isinstance(raw_id,(bytes,bytearray)) else str(raw_id)
            raw=client.get(f'{_REDIS_EVENT_PREFIX}:event:{eid}')
            if not raw: continue
            try: row=json.loads(raw.decode('utf-8') if isinstance(raw,(bytes,bytearray)) else raw)
            except Exception: continue
            if _event_local_upsert_v268(row): loaded+=1
        return loaded
    except Exception: return loaded

def _event_mark_mirrored_v268(event_ids,state_token=''):
    ids=[str(x) for x in (event_ids or []) if str(x)]
    if not ids: return 0
    n=0
    for eid in ids:
        row={'event_id':eid,'update_id':eid,'state':'mirrored','mirrored_at':time.time(),'state_token':str(state_token or '')[:120],'last_error':''}
        if _event_local_upsert_v268(row): n+=1
        _event_redis_store_v268(row)
    with STATE_LOCK:
        STATE['event_mirrored']=int(STATE.get('event_mirrored') or 0)+n; STATE['event_last_at']=time.time()
        try:
            conn=sqlite3.connect(EVENT_DB); STATE['event_pending']=int(conn.execute("SELECT COUNT(*) FROM events WHERE state IN ('received','failed_retry','committed')").fetchone()[0]); conn.close()
        except Exception: pass
    return n

def _event_pending_rows_v268(limit=100):
    _event_hydrate_pending_from_redis_v268(limit*2)
    _event_db_init_v268()
    with EVENT_LOCK:
        conn=sqlite3.connect(EVENT_DB,timeout=10)
        try:
            rows=conn.execute("SELECT event_id,update_id,chat_id,update_type,payload_json,payload_sha256,state,received_at,committed_at,state_token,last_error FROM events WHERE state IN ('received','failed_retry','committed') ORDER BY received_at ASC LIMIT ?",(max(1,min(250,int(limit))),)).fetchall()
        finally: conn.close()
    out=[]
    for r in rows:
        try: payload=json.loads(r[4] or '{}')
        except Exception: payload={}
        chat=None
        try: chat=int(r[2]) if str(r[2]).strip() else None
        except Exception: chat=r[2]
        out.append({'event_id':r[0],'update_id':r[1],'chat_id':chat,'update_type':r[3],'payload':payload,'payload_sha256':r[5],'state':r[6],'received_at':r[7],'committed_at':r[8],'state_token':r[9],'last_error':r[10]})
    with STATE_LOCK: STATE['event_pending']=len(out)
    return out

def _event_reconcile_loop_v268():
    while True:
        time.sleep(env_int('WORKER_EVENT_REDIS_RECONCILE_SEC',5,1,300))
        try:
            _event_hydrate_pending_from_redis_v268(500)
            # R15: any Worker-local event not yet guaranteed in Redis is retried here.
            _event_db_init_v268()
            with EVENT_LOCK:
                conn=sqlite3.connect(EVENT_DB,timeout=10)
                rows=conn.execute("SELECT event_id,update_id,chat_id,update_type,payload_json,payload_sha256,state,received_at,committed_at,state_token,last_error FROM events WHERE state IN ('received','failed_retry','committed') ORDER BY updated_at ASC LIMIT 200").fetchall()
                conn.close()
            for r in rows:
                try: payload=json.loads(r[4] or '{}')
                except Exception: payload={}
                row={'event_id':r[0],'update_id':r[1],'chat_id':r[2],'update_type':r[3],'payload':payload,'payload_sha256':r[5],'state':r[6],'received_at':r[7],'committed_at':r[8],'state_token':r[9],'last_error':r[10]}
                _event_redis_store_v268(row)
            # prune old mirrored local diagnostics; Redis keys expire independently.
            cutoff=time.time()-env_int('WORKER_EVENT_RETENTION_SEC',604800,86400,2592000)
            with EVENT_LOCK:
                conn=sqlite3.connect(EVENT_DB,timeout=10); conn.execute("DELETE FROM events WHERE state='mirrored' AND updated_at<?",(cutoff,)); conn.commit(); conn.close()
        except Exception as exc:
            with STATE_LOCK: STATE['event_last_error']=f'{type(exc).__name__}: {str(exc)[:180]}'

def _redis_client():
    global _REDIS_CLIENT
    if _redis is None:
        return None
    url = str(os.getenv('REDIS_URL','') or '').strip()
    if not url:
        return None
    with _REDIS_LOCK:
        if _REDIS_CLIENT is None:
            _REDIS_CLIENT = _redis.Redis.from_url(url, socket_connect_timeout=5, socket_timeout=12, health_check_interval=30)
        return _REDIS_CLIENT

def redis_store_snapshot(local_gz: Path, meta: dict, clear_deltas: bool=False):
    client = _redis_client()
    if client is None:
        return False, 'REDIS_URL not configured'
    try:
        payload = local_gz.read_bytes()
        max_bytes = env_int('WORKER_REDIS_SNAPSHOT_MAX_MB',16,1,128) * 1024 * 1024
        if len(payload) > max_bytes:
            return False, f'snapshot too large for Redis: {len(payload)} > {max_bytes}'
        incoming_revision = float((meta or {}).get('revision') or 0.0)
        existing_revision = 0.0
        try:
            existing_raw = client.get(_REDIS_META_KEY)
            if isinstance(existing_raw, (bytes, bytearray)):
                existing_raw = existing_raw.decode('utf-8', 'replace')
            existing_meta = json.loads(existing_raw) if isinstance(existing_raw, str) and existing_raw else {}
            existing_revision = float((existing_meta or {}).get('revision') or 0.0)
        except Exception:
            existing_revision = 0.0
        if existing_revision > incoming_revision + 0.000001:
            with STATE_LOCK:
                STATE['redis_cache_ok'] = True; STATE['redis_last_write'] = time.time(); STATE['redis_last_error'] = 'newer Redis snapshot preserved'
            return True, f'Redis newer snapshot preserved existing={existing_revision} incoming={incoming_revision}'
        row = {
            'revision': incoming_revision,
            'sha256_gz': str((meta or {}).get('sha256_gz') or hashlib.sha256(payload).hexdigest()),
            'size': len(payload), 'saved_at': time.time(), 'version': TRANSPORT_VERSION,
        }
        pipe = client.pipeline(transaction=True)
        pipe.set(_REDIS_SNAPSHOT_KEY, payload)
        pipe.set(_REDIS_META_KEY, json.dumps(row, separators=(',',':')))
        if clear_deltas:
            pipe.delete(_REDIS_DELTA_KEY)
            pipe.delete(_REDIS_DELTA_META_KEY)
        pipe.execute()
        with STATE_LOCK:
            STATE['redis_cache_ok'] = True; STATE['redis_last_write'] = time.time(); STATE['redis_last_error'] = ''
        return True, 'Redis snapshot cached'
    except Exception as exc:
        with STATE_LOCK:
            STATE['redis_cache_ok'] = False; STATE['redis_last_error'] = f'{type(exc).__name__}: {str(exc)[:220]}'
        return False, STATE['redis_last_error']

def redis_load_snapshot_to_cache():
    client = _redis_client()
    if client is None:
        return False, 'REDIS_URL not configured'
    fd, name = tempfile.mkstemp(prefix='vys262_redis_', suffix='.sqlite3.gz'); os.close(fd)
    tmp = Path(name)
    try:
        payload = client.get(_REDIS_SNAPSHOT_KEY)
        if not payload:
            return False, 'Redis snapshot missing'
        tmp.write_bytes(payload)
        ok, detail, meta = quick_check_gzip(tmp)
        if not ok:
            return False, 'Redis snapshot invalid: ' + str(detail)[:240]
        incoming_revision = float(meta.get('revision') or 0.0)
        with STATE_LOCK:
            current_revision = float(STATE.get('cache_revision') or 0.0)
        cache_exists = CACHE_LATEST.exists()
        accept = (not cache_exists) or current_revision <= 0.0 or (incoming_revision > 0.0 and incoming_revision >= current_revision)
        if accept:
            tmp_cache = CACHE_DIR / f'.redis_{secrets.token_hex(6)}.tmp'
            shutil.copy2(tmp, tmp_cache); os.replace(tmp_cache, CACHE_LATEST)
            with STATE_LOCK:
                STATE['cache_revision'] = max(current_revision, incoming_revision)
                STATE['last_snapshot_sha256'] = str(meta.get('sha256_gz') or '')
                STATE['last_snapshot_size'] = int(meta.get('size') or len(payload))
        elif cache_exists:
            with STATE_LOCK:
                STATE['redis_cache_ok'] = True; STATE['redis_last_read'] = time.time(); STATE['redis_last_error'] = 'older Redis snapshot ignored'
            return True, f'Redis snapshot older than live cache; kept cache revision={current_revision}'
        with STATE_LOCK:
            STATE['redis_cache_ok'] = True; STATE['redis_last_read'] = time.time(); STATE['redis_last_error'] = ''
        return True, 'Redis latest OK'
    except Exception as exc:
        with STATE_LOCK:
            STATE['redis_cache_ok'] = False; STATE['redis_last_error'] = f'{type(exc).__name__}: {str(exc)[:220]}'
        return False, STATE['redis_last_error']
    finally:
        tmp.unlink(missing_ok=True)


def run_cmd(args, timeout=120):
    return subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=timeout, check=False)


def mega_login():
    login_timeout = env_int('MEGA_LOGIN_TIMEOUT',120,30,300)
    try:
        who = run_cmd(['mega-whoami'], timeout=min(20,login_timeout))
        if who.returncode == 0: return True, 'already logged in'
    except FileNotFoundError: return False, 'MEGAcmd not installed'
    except Exception: pass
    session = str(os.getenv('MEGA_SESSION','') or '').strip(); email = str(os.getenv('MEGA_EMAIL','') or '').strip(); password = str(os.getenv('MEGA_PASSWORD','') or '').strip()
    if not session and (not email or not password): return False, 'MEGA_SESSION or MEGA_EMAIL/MEGA_PASSWORD missing'
    args = ['mega-login', session] if session else ['mega-login', email, password]
    try: run_cmd(['mega-logout'], timeout=20)
    except Exception: pass
    for attempt in (1,2):
        try:
            p = run_cmd(args, timeout=login_timeout)
            if p.returncode == 0: return True, 'login OK'
            err = (p.stderr or p.stdout or 'mega-login rejected')[:220]
        except subprocess.TimeoutExpired: return False, f'mega-login timeout after {login_timeout}s'
        except Exception as exc: return False, f'mega-login {type(exc).__name__}'
        if attempt == 1:
            try: run_cmd(['mega-logout'], timeout=20)
            except Exception: pass
            time.sleep(.8)
    return False, err


def mega_exists(path):
    try: return run_cmd(['mega-ls', path], timeout=30).returncode == 0
    except Exception: return False


def ensure_mega_dir(path):
    if mega_exists(path): return True
    try: p = run_cmd(['mega-mkdir', path], timeout=45)
    except Exception: return False
    return p.returncode == 0 or mega_exists(path)


def prepare_mega_layout():
    ok, detail = mega_login()
    if not ok: return False, detail
    for p in (mega_root(), remote_db_dir(), remote_history_dir()):
        if not ensure_mega_dir(p): return False, f'cannot create MEGA dir {p}'
    with STATE_LOCK: STATE['mega_layout_ok'] = True
    return True, 'MEGA layout ready'


def mega_legacy_roots():
    raw = str(os.getenv('MEGA_LEGACY_BACKUP_DIRS','/TelegramBotBackups-2T,/TelegramBotBackups') or '')
    out = []
    current = mega_root()
    for item in raw.split(','):
        item = '/' + str(item or '').strip().strip('/')
        if item != '/' and item != current and item not in out:
            out.append(item)
    return out


def mega_restore_candidates():
    rows = [remote_latest()]
    for root in mega_legacy_roots():
        rows.append(root.rstrip('/') + '/database/latest_bot_state.sqlite3.gz')
    return rows


def quick_check_gzip(gz_path: Path):
    work = Path(tempfile.mkdtemp(prefix='vys262_check_')); raw = work / 'state.sqlite3'
    try:
        with gzip.open(gz_path,'rb') as src, open(raw,'wb') as dst: shutil.copyfileobj(src,dst,1024*1024)
        con = sqlite3.connect(str(raw))
        continuity_revision = 0.0
        try:
            row = con.execute('PRAGMA quick_check').fetchone()
            user_state_seq = 0
            try:
                for kind in ('split_state_revision_r18','runtime_continuity_v263','user_state_shadow_v265'):
                    meta_row = con.execute("SELECT v FROM meta WHERE kind=? AND k=?", (kind,'latest')).fetchone()
                    if not meta_row:
                        continue
                    meta_payload = json.loads(meta_row[0]) if isinstance(meta_row[0], str) else (meta_row[0] or {})
                    continuity_revision = max(continuity_revision, float((meta_payload or {}).get('saved_at') or 0.0))
                    if kind == 'user_state_shadow_v265':
                        user_state_seq = int((meta_payload or {}).get('seq') or 0)
            except Exception:
                pass
        finally: con.close()
        if not row or str(row[0]).lower() != 'ok': return False, f'quick_check={row}', {}
        payload = gz_path.read_bytes(); return True, 'OK', {'sha256_gz':hashlib.sha256(payload).hexdigest(),'size':len(payload),'revision':continuity_revision,'user_state_seq':user_state_seq}
    except Exception as exc: return False, f'{type(exc).__name__}: {str(exc)[:180]}', {}
    finally: shutil.rmtree(work, ignore_errors=True)



DELTA_APPLY_LOCK = threading.RLock()

def _sha256_file_v267(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()

def _sqlite_page_size_v267(path: Path) -> int:
    with open(path, 'rb') as fh:
        head = fh.read(100)
    if len(head) < 18 or head[:16] != b'SQLite format 3\x00':
        raise RuntimeError('not a SQLite database')
    size = int.from_bytes(head[16:18], 'big')
    if size == 1:
        size = 65536
    if size < 512 or size > 65536 or (size & (size - 1)):
        raise RuntimeError(f'invalid SQLite page size {size}')
    return size

def _quick_check_db_v267(path: Path):
    try:
        con = sqlite3.connect(str(path))
        try:
            row = con.execute('PRAGMA quick_check').fetchone()
            revision = 0.0
            user_state_seq = 0
            try:
                for kind in ('split_state_revision_r18','runtime_continuity_v263','user_state_shadow_v265'):
                    meta_row = con.execute("SELECT v FROM meta WHERE kind=? AND k=?", (kind,'latest')).fetchone()
                    if not meta_row:
                        continue
                    payload = json.loads(meta_row[0]) if isinstance(meta_row[0], str) else (meta_row[0] or {})
                    revision = max(revision, float((payload or {}).get('saved_at') or 0.0))
                    if kind == 'user_state_shadow_v265':
                        user_state_seq = int((payload or {}).get('seq') or 0)
            except Exception:
                pass
        finally:
            con.close()
        if not row or str(row[0]).lower() != 'ok':
            return False, f'quick_check={row}', {}
        return True, 'OK', {'revision':revision,'user_state_seq':user_state_seq,'sha256_db':_sha256_file_v267(path),'size_db':path.stat().st_size}
    except Exception as exc:
        return False, f'{type(exc).__name__}: {str(exc)[:220]}', {}

def _gzip_cache_db_v267():
    if not CACHE_DB.exists():
        return False, 'CACHE_DB missing', {}
    fd, name = tempfile.mkstemp(prefix='vys262_cache_', suffix='.sqlite3.gz'); os.close(fd)
    tmp = Path(name)
    try:
        with open(CACHE_DB,'rb') as src, gzip.open(tmp,'wb',compresslevel=9) as dst:
            shutil.copyfileobj(src,dst,1024*1024)
        ok, detail, meta = quick_check_gzip(tmp)
        if not ok:
            return False, detail, {}
        target = CACHE_DIR / f'.latest_gz_{secrets.token_hex(6)}.tmp'
        shutil.copy2(tmp,target); os.replace(target,CACHE_LATEST)
        return True, 'OK', meta
    finally:
        tmp.unlink(missing_ok=True)

def _ensure_cache_db_v267():
    if CACHE_DB.exists():
        ok, detail, meta = _quick_check_db_v267(CACHE_DB)
        if ok:
            with STATE_LOCK:
                STATE['last_state_sha256'] = str(meta.get('sha256_db') or '')
            return True, 'cache DB OK', meta
        CACHE_DB.unlink(missing_ok=True)
    if not CACHE_LATEST.exists():
        return False, 'CACHE_LATEST missing', {}
    fd, name = tempfile.mkstemp(prefix='vys262_cache_db_', suffix='.sqlite3'); os.close(fd)
    tmp = Path(name)
    try:
        with gzip.open(CACHE_LATEST,'rb') as src, open(tmp,'wb') as dst:
            shutil.copyfileobj(src,dst,1024*1024)
        ok, detail, meta = _quick_check_db_v267(tmp)
        if not ok:
            return False, 'decompressed cache invalid: '+detail, {}
        target = CACHE_DIR / f'.latest_db_{secrets.token_hex(6)}.tmp'
        shutil.copy2(tmp,target); os.replace(target,CACHE_DB)
        with STATE_LOCK:
            STATE['last_state_sha256'] = str(meta.get('sha256_db') or '')
        return True, 'cache DB hydrated', meta
    finally:
        tmp.unlink(missing_ok=True)

def _redis_append_delta_v267(wire: bytes, payload: dict):
    client = _redis_client()
    if client is None:
        return False, 'REDIS_URL not configured'
    try:
        max_items = env_int('WORKER_REDIS_DELTA_MAX_ITEMS',2000,10,20000)
        row = {'new_sha256':str(payload.get('new_sha256') or ''), 'state_token':str(payload.get('state_token') or ''), 'saved_at':time.time(), 'count':int(STATE.get('delta_since_checkpoint') or 0)}
        pipe=client.pipeline(transaction=True)
        pipe.rpush(_REDIS_DELTA_KEY, wire)
        pipe.ltrim(_REDIS_DELTA_KEY, -max_items, -1)
        pipe.set(_REDIS_DELTA_META_KEY, json.dumps(row,separators=(',',':')))
        pipe.execute()
        with STATE_LOCK:
            STATE['redis_cache_ok']=True; STATE['redis_last_write']=time.time(); STATE['redis_last_error']=''
        return True, 'Redis delta appended'
    except Exception as exc:
        detail=f'{type(exc).__name__}: {str(exc)[:220]}'
        with STATE_LOCK:
            STATE['redis_cache_ok']=False; STATE['redis_last_error']=detail
        return False,detail

def _apply_delta_payload_v267(payload: dict, *, journal_wire: bytes|None=None, replay=False):
    if not isinstance(payload,dict) or int(payload.get('schema') or 0) != 1:
        return False, 'invalid delta schema', 'invalid'
    ok, detail, meta = _ensure_cache_db_v267()
    if not ok:
        return False, detail, 'need_full'
    base_sha=str(payload.get('base_sha256') or '').lower()
    new_sha=str(payload.get('new_sha256') or '').lower()
    current_sha=str(meta.get('sha256_db') or _sha256_file_v267(CACHE_DB)).lower()
    if new_sha and current_sha == new_sha:
        _event_mark_mirrored_v268(payload.get('event_ids') or [], payload.get('state_token') or '')
        return True, 'already applied', 'up_to_date'
    if not base_sha or current_sha != base_sha:
        return False, f'base mismatch worker={current_sha[:16]} front={base_sha[:16]}', 'need_full'
    page_size=int(payload.get('page_size') or 0)
    db_size=int(payload.get('db_size') or 0)
    pages=payload.get('pages') or []
    if page_size != _sqlite_page_size_v267(CACHE_DB) or db_size < 100 or db_size > env_int('WORKER_DELTA_MAX_DB_MB',128,1,1024)*1024*1024:
        return False, 'delta database geometry invalid', 'need_full'
    if not isinstance(pages,list) or len(pages) > env_int('WORKER_DELTA_MAX_PAGES',4096,8,65536):
        return False, 'too many delta pages', 'invalid'
    tmp=CACHE_DIR / f'.delta_{secrets.token_hex(8)}.sqlite3'
    shutil.copy2(CACHE_DB,tmp)
    try:
        with open(tmp,'r+b') as fh:
            for row in pages:
                if not isinstance(row,list) or len(row)!=2:
                    return False,'invalid delta page row','invalid'
                idx=int(row[0]); data=base64.b64decode(str(row[1] or ''),validate=True)
                if idx < 0 or len(data) > page_size:
                    return False,'invalid delta page geometry','invalid'
                fh.seek(idx*page_size); fh.write(data)
            fh.truncate(db_size)
            fh.flush(); os.fsync(fh.fileno())
        ok2, detail2, meta2=_quick_check_db_v267(tmp)
        if not ok2:
            return False,'patched DB invalid: '+detail2,'need_full'
        actual=str(meta2.get('sha256_db') or '').lower()
        if new_sha and actual != new_sha:
            return False,f'patched sha mismatch expected={new_sha[:16]} actual={actual[:16]}','need_full'
        # Durability first: after the patch is fully validated, append the tiny delta to
        # Redis BEFORE replacing the local /tmp cache. If the Worker is killed in the
        # next millisecond, startup can replay this journal onto the last checkpoint.
        if (not replay) and journal_wire is not None and _redis_client() is not None:
            redis_ok,redis_detail=_redis_append_delta_v267(journal_wire,payload)
            if not redis_ok:
                return False,'Redis delta append failed: '+str(redis_detail)[:180],'redis_unavailable'
        os.replace(tmp,CACHE_DB)
        gz_ok,gz_detail,gz_meta=_gzip_cache_db_v267()
        if not gz_ok:
            return False,'cache gzip failed: '+gz_detail,'need_full'
        with STATE_LOCK:
            STATE['cache_revision']=max(float(STATE.get('cache_revision') or 0.0),float(meta2.get('revision') or 0.0))
            STATE['last_snapshot_sha256']=str(gz_meta.get('sha256_gz') or '')
            STATE['last_snapshot_size']=int(gz_meta.get('size') or 0)
            STATE['last_snapshot_at']=time.time()
            STATE['last_state_sha256']=actual
            STATE['last_state_token']=str(payload.get('state_token') or '')[:120]
            STATE['delta_last_at']=time.time(); STATE['delta_last_pages']=len(pages); STATE['delta_last_error']=''
            if replay:
                STATE['delta_replayed']=int(STATE.get('delta_replayed') or 0)+1
            else:
                STATE['delta_count']=int(STATE.get('delta_count') or 0)+1
                STATE['delta_since_checkpoint']=int(STATE.get('delta_since_checkpoint') or 0)+1
                STATE['delta_bytes']=int(STATE.get('delta_bytes') or 0)+int(len(journal_wire or b''))
                STATE['delta_bytes_since_checkpoint']=int(STATE.get('delta_bytes_since_checkpoint') or 0)+int(len(journal_wire or b''))
        _event_mark_mirrored_v268(payload.get('event_ids') or [], payload.get('state_token') or '')
        return True,f'applied pages={len(pages)} bytes={len(journal_wire or b"")}', 'applied'
    finally:
        tmp.unlink(missing_ok=True)

def redis_replay_deltas_v267():
    client=_redis_client()
    if client is None:
        return False,'REDIS_URL not configured'
    try:
        rows=client.lrange(_REDIS_DELTA_KEY,0,-1) or []
        if not rows:
            _ensure_cache_db_v267()
            return True,'no Redis deltas'
        applied=0
        for wire in rows:
            try:
                raw=gzip.decompress(wire)
                payload=json.loads(raw.decode('utf-8'))
            except Exception as exc:
                return False,f'delta decode failed at {applied}: {exc}'
            ok,detail,status=_apply_delta_payload_v267(payload,journal_wire=None,replay=True)
            if not ok and status!='up_to_date':
                return False,f'delta replay failed at {applied}: {detail}'
            applied+=1
        with STATE_LOCK:
            STATE['delta_since_checkpoint']=max(int(STATE.get('delta_since_checkpoint') or 0), applied)
        return True,f'replayed {applied} Redis deltas'
    except Exception as exc:
        return False,f'{type(exc).__name__}: {str(exc)[:220]}'

def _full_checkpoint_v267(reason='periodic'):
    with DELTA_APPLY_LOCK:
        ok,detail,meta=_ensure_cache_db_v267()
        if not ok:
            return False,detail
        gz_ok,gz_detail,gz_meta=_gzip_cache_db_v267()
        if not gz_ok:
            return False,gz_detail
        redis_ok,redis_detail=redis_store_snapshot(CACHE_LATEST,gz_meta,clear_deltas=True)
        with STATE_LOCK: last_mega=float(STATE.get('last_mega_upload_at') or 0.0)
        mega_every=env_int('WORKER_MEGA_CHECKPOINT_SEC',86400,3600,604800)
        mega_due=(last_mega<=0.0 or time.time()-last_mega>=mega_every or str(reason).startswith(('manual','shutdown','reconcile')))
        if mega_due:
            mega_ok,mega_detail=mega_promote_snapshot(CACHE_LATEST)
        else:
            mega_ok,mega_detail=True,f'deferred until {mega_every}s interval'
        with STATE_LOCK:
            if mega_due and mega_ok: STATE['last_mega_upload_at']=time.time()
            STATE['full_checkpoint_at']=time.time(); STATE['full_checkpoint_count']=int(STATE.get('full_checkpoint_count') or 0)+1
            if redis_ok:
                STATE['delta_since_checkpoint']=0; STATE['delta_bytes_since_checkpoint']=0
        return bool(redis_ok),f'{reason}: redis={redis_ok} {redis_detail}; mega={mega_ok} {mega_detail}'

def _checkpoint_loop_v267():
    while True:
        time.sleep(30)
        try:
            with STATE_LOCK:
                count=int(STATE.get('delta_since_checkpoint') or 0)
                last=float(STATE.get('full_checkpoint_at') or STATE.get('started_at') or time.time())
            if count <= 0:
                continue
            seconds=env_int('WORKER_FULL_CHECKPOINT_SEC',21600,300,86400)
            max_deltas=env_int('WORKER_FULL_CHECKPOINT_MAX_DELTAS',1000,10,20000)
            with STATE_LOCK: delta_bytes_since=int(STATE.get('delta_bytes_since_checkpoint') or 0)
            max_delta_bytes=env_int('WORKER_FULL_CHECKPOINT_MAX_DELTA_MB',16,1,128)*1024*1024
            if count >= max_deltas or delta_bytes_since >= max_delta_bytes or time.time()-last >= seconds:
                ok,detail=_full_checkpoint_v267('threshold')
                print(f'[R12 CHECKPOINT] ok={ok} {detail}',flush=True)
        except Exception as exc:
            print(f'[R12 CHECKPOINT ERROR] {type(exc).__name__}: {str(exc)[:240]}',flush=True)


def fetch_front_snapshot():
    base, secret = front_base(), peer_secret()
    if not base or not secret: return None, 'front URL/secret not configured', {}
    timeout = env_int('WORKER_FRONT_FETCH_TIMEOUT',30,5,180)
    work = Path(tempfile.mkdtemp(prefix='vys262_front_fetch_')); path = work / 'snapshot.sqlite3.gz'
    try:
        r = requests.get(base + '/internal/split/state', headers={'X-Peer-Secret':secret,'User-Agent':'vys-262-worker-fetch-r19','X-Split-Rebase':'1'}, timeout=timeout, stream=True)
        if r.status_code != 200: return None, f'front HTTP {r.status_code}: {r.text[:180]}', {}
        with open(path,'wb') as fh:
            for chunk in r.iter_content(1024*1024):
                if chunk: fh.write(chunk)
        ok, detail, meta = quick_check_gzip(path)
        if not ok: return None, detail, meta
        meta['state_token'] = str(r.headers.get('X-Split-State-Token') or '')[:120]
        # Caller owns the copied temp path, so move it out of disposable dir.
        fd, durable_name = tempfile.mkstemp(prefix='vys262_snapshot_', suffix='.sqlite3.gz'); os.close(fd)
        durable = Path(durable_name); shutil.copy2(path, durable)
        return durable, 'front snapshot OK', meta
    except Exception as exc: return None, f'front {type(exc).__name__}: {str(exc)[:180]}', {}
    finally: shutil.rmtree(work, ignore_errors=True)


def mega_promote_snapshot(local_gz: Path):
    with MEGA_LOCK:
        ok, detail = prepare_mega_layout()
        if not ok: return False, detail
        stamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        # Keep the old latest as history before replacement. Failure to archive does not delete latest.
        if mega_exists(remote_latest()):
            hist_tmp = remote_history_dir().rstrip('/') + '/latest_bot_state.sqlite3.gz'
            try:
                cp = run_cmd(['mega-cp', remote_latest(), remote_history_dir()], timeout=120)
                if cp.returncode == 0 and mega_exists(hist_tmp):
                    run_cmd(['mega-mv', hist_tmp, remote_history_dir().rstrip('/') + f'/state_{stamp}.sqlite3.gz'], timeout=60)
            except Exception: pass
        work = Path(tempfile.mkdtemp(prefix='vys262_mega_put_'))
        try:
            incoming_name = f'incoming_{stamp}_{secrets.token_hex(4)}.sqlite3.gz'; incoming_local = work / incoming_name; shutil.copy2(local_gz, incoming_local)
            put = run_cmd(['mega-put', str(incoming_local), remote_db_dir()], timeout=env_int('MEGA_TIMEOUT',180,30,900))
            if put.returncode != 0: return False, 'mega-put failed: ' + (put.stderr or put.stdout or '')[:180]
            incoming_remote = remote_db_dir().rstrip('/') + '/' + incoming_name
            rollback_remote = ''
            if mega_exists(remote_latest()):
                rollback_remote = remote_history_dir().rstrip('/') + f'/rollback_{stamp}_{secrets.token_hex(3)}.sqlite3.gz'
                old_mv = run_cmd(['mega-mv', remote_latest(), rollback_remote], timeout=90)
                if old_mv.returncode != 0 and mega_exists(remote_latest()):
                    return False, 'could not move old latest to rollback slot'
            mv = run_cmd(['mega-mv', incoming_remote, remote_latest()], timeout=90)
            if mv.returncode != 0 and not mega_exists(remote_latest()):
                if rollback_remote and mega_exists(rollback_remote):
                    try: run_cmd(['mega-mv', rollback_remote, remote_latest()], timeout=90)
                    except Exception: pass
                return False, 'mega-mv promote failed; rollback attempted'
            # The rollback object is useful history; rename it to a normal timestamped history name.
            if rollback_remote and mega_exists(rollback_remote):
                try: run_cmd(['mega-mv', rollback_remote, remote_history_dir().rstrip('/') + f'/state_{stamp}_previous.sqlite3.gz'], timeout=60)
                except Exception: pass
            return True, 'MEGA latest promoted'
        finally: shutil.rmtree(work, ignore_errors=True)


def sync_state_job(job):
    with STATE_LOCK:
        STATE['last_full_fetch_started'] = time.time()
    snap, detail, meta = fetch_front_snapshot()
    if not snap:
        return False, detail
    try:
        incoming_revision = float(meta.get('revision') or 0.0)
        with STATE_LOCK:
            current_revision = float(STATE.get('cache_revision') or 0.0)
        cache_exists = CACHE_LATEST.exists()
        # Never let a delayed GET overwrite a newer direct-shutdown upload/cache.
        if cache_exists and current_revision > 0.0 and (incoming_revision <= 0.0 or incoming_revision < current_revision):
            fetched_token = str(meta.get('state_token') or job.get('state_token') or '')[:120]
            with STATE_LOCK:
                if fetched_token: STATE['last_state_token'] = fetched_token
                STATE['deduped_sync_requests'] += 1
            return True, f'stale snapshot ignored revision={incoming_revision} < cache={current_revision}'
        # R6 durability order: validated front state becomes the restore source FIRST.
        tmp_cache = CACHE_DIR / f'.sync_{secrets.token_hex(6)}.tmp'
        shutil.copy2(snap, tmp_cache)
        os.replace(tmp_cache, CACHE_LATEST)
        with STATE_LOCK:
            STATE['cache_revision'] = max(current_revision, incoming_revision)
            STATE['last_snapshot_sha256'] = str(meta.get('sha256_gz') or '')
            STATE['last_snapshot_size'] = int(meta.get('size') or 0)
            STATE['last_snapshot_at'] = time.time()
            fetched_token = str(meta.get('state_token') or job.get('state_token') or '')[:120]
            if fetched_token:
                STATE['last_state_token'] = fetched_token
        # Full resync establishes a fresh delta base and clears the Redis delta journal.
        try:
            CACHE_DB.unlink(missing_ok=True)
            _ensure_cache_db_v267()
            with STATE_LOCK:
                STATE['last_state_sha256'] = _sha256_file_v267(CACHE_DB) if CACHE_DB.exists() else ''
                STATE['delta_since_checkpoint'] = 0
                STATE['delta_bytes_since_checkpoint'] = 0
        except Exception:
            pass
        # Fast durable cache is written before slow archival MEGA.
        redis_ok, redis_detail = redis_store_snapshot(CACHE_LATEST, meta, clear_deltas=True)
        mega_ok, mega_detail = mega_promote_snapshot(CACHE_LATEST)
        with STATE_LOCK:
            if mega_ok:
                STATE['last_mega_upload_at'] = time.time()
            else:
                STATE['sync_failures'] += 1
            STATE['sync_count'] += 1
        return True, f"cached {meta.get('size',0)} bytes; redis={redis_ok}; mega={mega_ok}; {redis_detail}; {mega_detail}"
    finally:
        snap.unlink(missing_ok=True)


def _download_mega_latest():
    with MEGA_LOCK:
        ok, detail = prepare_mega_layout()
        if not ok:
            return None, detail
        last_detail = 'MEGA latest snapshot not found'
        for idx, remote in enumerate(mega_restore_candidates()):
            if not mega_exists(remote):
                continue
            work = Path(tempfile.mkdtemp(prefix=f'vys262_mega_get_{idx}_'))
            try:
                p = run_cmd(['mega-get', remote, str(work)], timeout=env_int('MEGA_TIMEOUT',180,30,900))
                if p.returncode != 0:
                    last_detail = 'mega-get failed: ' + (p.stderr or p.stdout or '')[:180]
                    continue
                rows = list(work.rglob('latest_bot_state.sqlite3.gz')) + list(work.rglob('*.sqlite3.gz'))
                if not rows:
                    last_detail = 'MEGA download has no SQLite gzip'
                    continue
                check_ok, check_detail, meta = quick_check_gzip(rows[0])
                if not check_ok:
                    last_detail = check_detail
                    continue
                incoming_revision = float(meta.get('revision') or 0.0)
                with STATE_LOCK:
                    current_revision = float(STATE.get('cache_revision') or 0.0)
                cache_exists = CACHE_LATEST.exists()
                accept = (not cache_exists) or current_revision <= 0.0 or (incoming_revision > 0.0 and incoming_revision >= current_revision)
                if accept:
                    tmp_cache = CACHE_DIR / f'.mega_{secrets.token_hex(6)}.tmp'
                    shutil.copy2(rows[0], tmp_cache); os.replace(tmp_cache, CACHE_LATEST)
                    with STATE_LOCK:
                        STATE['cache_revision'] = max(current_revision, incoming_revision)
                        STATE['last_snapshot_sha256'] = str(meta.get('sha256_gz') or '')
                        STATE['last_snapshot_size'] = int(meta.get('size') or 0)
                    try:
                        CACHE_DB.unlink(missing_ok=True)
                        _ensure_cache_db_v267()
                        redis_store_snapshot(CACHE_LATEST, meta)
                    except Exception:
                        pass
                with STATE_LOCK:
                    STATE['last_restore_download_at'] = time.time()
                if not accept and cache_exists:
                    return CACHE_LATEST, f'MEGA snapshot older than restore cache; kept cache rev={current_revision}'
                return CACHE_LATEST, 'MEGA latest OK' if idx == 0 else f'MEGA legacy restore cache OK: {remote}'
            finally:
                shutil.rmtree(work, ignore_errors=True)
        return None, last_detail


def process_job(job):
    global SYNC_PENDING, SYNC_DIRTY, SYNC_DIRTY_REASON
    jid, kind = str(job.get('id') or ''), str(job.get('type') or '')
    with STATE_LOCK:
        STATE['job_last_id']=jid; STATE['job_last_type']=kind; STATE['job_last_reason']=str(job.get('reason') or '')[:180]; STATE['job_last_started']=time.time(); STATE['job_last_error']=''
    try:
        if kind == 'sync_state':
            ok, detail = sync_state_job(job)
        elif kind == 'promote_uploaded':
            path = Path(str(job.get('path') or ''))
            try:
                if not path.exists():
                    ok, detail = False, 'uploaded snapshot missing before promote'
                else:
                    ok, detail = mega_promote_snapshot(path)
                    if ok:
                        with STATE_LOCK:
                            STATE['last_mega_upload_at'] = time.time()
                            STATE['sync_count'] += 1
            finally:
                try: path.unlink(missing_ok=True)
                except Exception: pass
        else:
            ok, detail = False, f'unsupported job type: {kind}'
        with STATE_LOCK:
            STATE['job_last_done']=time.time(); STATE['job_last_error']='' if ok else detail[:260]
            if not ok: STATE['sync_failures'] += 1
        print(f'[WORKER JOB] {jid} {kind} ok={ok} {detail}', flush=True)
    finally:
        if kind == 'sync_state':
            followup = None
            with STATE_LOCK:
                fetched_token = str(STATE.get('last_state_token') or '')
            with SYNC_PENDING_LOCK:
                dirty_token = str(STATE.get('dirty_state_token') or '')
                # A duplicate request that arrived while the first GET was running must
                # not trigger a second full SQLite download. Only a genuinely newer
                # front state token deserves one follow-up fetch.
                if SYNC_DIRTY and dirty_token and dirty_token != fetched_token:
                    # R26: never chain another full Front snapshot just because state
                    # changed while this full snapshot was in flight. FAST will send
                    # compact deltas from the newly promoted baseline. If those deltas
                    # truly diverge, the Front schedules one idle full reconcile.
                    with STATE_LOCK:
                        STATE['full_fetch_suppressed'] = int(STATE.get('full_fetch_suppressed') or 0) + 1
                        STATE['last_full_suppressed_token'] = dirty_token
                    SYNC_DIRTY = False
                    SYNC_DIRTY_REASON = ''
                    SYNC_PENDING = False
                    with STATE_LOCK:
                        STATE['active_state_token'] = ''
                        STATE['dirty_state_token'] = ''
                else:
                    SYNC_DIRTY = False
                    SYNC_DIRTY_REASON = ''
                    SYNC_PENDING = False
                    with STATE_LOCK:
                        STATE['active_state_token'] = ''
                        STATE['dirty_state_token'] = ''
            if followup is not None:
                try:
                    JOB_Q.put_nowait(followup)
                except queue.Full:
                    with SYNC_PENDING_LOCK:
                        SYNC_PENDING = False
                        SYNC_DIRTY = True
                        SYNC_DIRTY_REASON = str(followup.get('reason') or 'queue_full')
                    with STATE_LOCK:
                        STATE['dirty_state_token'] = str(followup.get('state_token') or '')[:120]


def worker_loop():
    while True:
        job = JOB_Q.get()
        try: process_job(job)
        except Exception as exc:
            with STATE_LOCK: STATE['job_last_error']=f'{type(exc).__name__}: {str(exc)[:240]}'; STATE['sync_failures'] += 1
        finally: JOB_Q.task_done()


def _restore_refresh_background():
    global RESTORE_REFRESH_RUNNING
    try: _download_mega_latest()
    finally:
        with RESTORE_REFRESH_LOCK: RESTORE_REFRESH_RUNNING = False

def _schedule_restore_refresh():
    global RESTORE_REFRESH_RUNNING
    with RESTORE_REFRESH_LOCK:
        if RESTORE_REFRESH_RUNNING: return False
        RESTORE_REFRESH_RUNNING = True
    threading.Thread(target=_restore_refresh_background, name='vys262-worker-restore-refresh', daemon=True).start(); return True


def _reconcile_hash_loop_v268():
    """Every ~6h compare hashes; transfer a full database only on real divergence."""
    while True:
        interval=env_int('WORKER_RECONCILE_SEC',21600,900,86400)
        time.sleep(interval)
        base,secret=front_base(),peer_secret()
        if not base or not secret: continue
        with STATE_LOCK: STATE['reconcile_last_at']=time.time()
        try:
            r=requests.get(base+'/internal/split/hash',headers={'X-Peer-Secret':secret,'User-Agent':'vys-262-worker-reconcile-r13'},timeout=30)
            if r.status_code!=200: raise RuntimeError(f'hash HTTP {r.status_code}: {r.text[:160]}')
            body=r.json() if r.content else {}; front_sha=str(body.get('sha256') or '')
            with STATE_LOCK: worker_sha=str(STATE.get('last_state_sha256') or '')
            if front_sha and worker_sha and front_sha==worker_sha:
                with STATE_LOCK: STATE['reconcile_last_ok']=time.time(); STATE['reconcile_last_error']=''
                print(f'[R13 RECONCILE] hash OK {front_sha[:16]}',flush=True); continue
            # Rare recovery: only now fetch the ~1 MB full compressed SQLite.
            ok,detail=sync_state_job({'state_token':str(body.get('state_token') or ''),'reason':'reconcile_hash_mismatch'})
            with STATE_LOCK:
                if ok:
                    STATE['reconcile_last_ok']=time.time(); STATE['reconcile_last_error']=''; STATE['reconcile_full_resyncs']=int(STATE.get('reconcile_full_resyncs') or 0)+1
                else: STATE['reconcile_last_error']=str(detail)[:220]
            print(f'[R13 RECONCILE] full={ok} {detail}',flush=True)
        except Exception as exc:
            with STATE_LOCK: STATE['reconcile_last_error']=f'{type(exc).__name__}: {str(exc)[:220]}'
            print(f'[R13 RECONCILE ERROR] {type(exc).__name__}: {str(exc)[:220]}',flush=True)

def peer_loop():
    time.sleep(5)
    while True:
        if not env_bool('PEER_PING_ENABLED', True):
            time.sleep(env_int('PEER_PING_INTERVAL_SEC',120,30,1800))
            continue
        base = front_base()
        with STATE_LOCK: STATE['peer_last_attempt'] = time.time()
        if base:
            try:
                headers={'User-Agent':'vys-262-worker-peer-r11'}
                if peer_secret(): headers['X-Peer-Secret']=peer_secret()
                r = requests.get(base + '/peer/health', headers=headers, timeout=12)
                payload = {}
                try: payload = r.json() if r.content else {}
                except Exception: payload = {}
                with STATE_LOCK:
                    STATE['peer_status']=int(r.status_code)
                    if 200 <= r.status_code < 300:
                        STATE['peer_last_ok']=time.time(); STATE['peer_last_error']=''
                        STATE['front_health']={
                            'ok':bool(payload.get('ok', True)), 'version':str(payload.get('version') or ''),
                            'bot_version':str(payload.get('bot_version') or ''), 'ready':bool(payload.get('ready', False)),
                            'phase':str(payload.get('phase') or ''), 'seen_at':time.time(),
                        }
                    else: STATE['peer_last_error']=f'HTTP {r.status_code}'
            except Exception as exc:
                with STATE_LOCK: STATE['peer_status']=None; STATE['peer_last_error']=str(exc)[:220]
        else:
            with STATE_LOCK: STATE['peer_last_error']='FRONT_SERVICE_URL empty'
        time.sleep(env_int('PEER_PING_INTERVAL_SEC',120,30,1800))


def _b64url(raw: bytes): return base64.urlsafe_b64encode(raw).rstrip(b'=').decode('ascii')
def _google_info():
    raw = str(os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON','') or '').strip()
    if not raw: raise RuntimeError('GOOGLE_SERVICE_ACCOUNT_JSON is not configured on worker')
    try: info = json.loads(raw) if raw.lstrip().startswith('{') else json.loads(base64.b64decode(raw).decode('utf-8'))
    except Exception as exc: raise RuntimeError(f'GOOGLE_SERVICE_ACCOUNT_JSON invalid: {exc}')
    for key in ('client_email','private_key','token_uri'):
        if not info.get(key): raise RuntimeError(f'GOOGLE_SERVICE_ACCOUNT_JSON missing {key}')
    return info

def _google_sign(message: bytes, private_key: str):
    key_path = msg_path = sig_path = None
    try:
        with tempfile.NamedTemporaryFile('w', encoding='utf-8', delete=False) as fh: fh.write(private_key); key_path = fh.name
        with tempfile.NamedTemporaryFile('wb', delete=False) as fh: fh.write(message); msg_path = fh.name
        fd, sig_path = tempfile.mkstemp(prefix='google_jwt_', suffix='.sig'); os.close(fd)
        p = subprocess.run(['openssl','dgst','-sha256','-sign',key_path,'-out',sig_path,msg_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=20)
        if p.returncode != 0: raise RuntimeError(p.stderr.decode('utf-8','replace')[-400:])
        return Path(sig_path).read_bytes()
    finally:
        for p in (key_path,msg_path,sig_path):
            if p:
                try: os.remove(p)
                except Exception: pass

def _google_token():
    with GOOGLE_LOCK:
        now = time.time()
        if _GOOGLE_TOKEN['token'] and now < float(_GOOGLE_TOKEN['expires_at']) - 120: return _GOOGLE_TOKEN['token']
        info = _google_info(); header={'alg':'RS256','typ':'JWT'}; claims={'iss':info['client_email'],'scope':'https://www.googleapis.com/auth/spreadsheets https://www.googleapis.com/auth/drive','aud':info.get('token_uri') or 'https://oauth2.googleapis.com/token','iat':int(now),'exp':int(now)+3600}
        signing = (_b64url(json.dumps(header,separators=(',',':')).encode()) + '.' + _b64url(json.dumps(claims,separators=(',',':')).encode())).encode('ascii')
        assertion = signing.decode('ascii') + '.' + _b64url(_google_sign(signing, info['private_key']))
        r = requests.post(info.get('token_uri') or 'https://oauth2.googleapis.com/token', data={'grant_type':'urn:ietf:params:oauth:grant-type:jwt-bearer','assertion':assertion}, timeout=30)
        if r.status_code >= 300: raise RuntimeError(f'Google OAuth {r.status_code}: {r.text[:400]}')
        payload = r.json(); token = str(payload.get('access_token') or '')
        if not token: raise RuntimeError('Google OAuth returned no access_token')
        _GOOGLE_TOKEN.update(token=token, expires_at=now+int(payload.get('expires_in',3600) or 3600)); return token

def _sheet_id(value=None):
    raw = str(value if value is not None and str(value).strip() else os.getenv('GOOGLE_SHEETS_SPREADSHEET_ID','') or '').strip()
    m = re.search(r'/spreadsheets/d/([A-Za-z0-9_-]+)', raw)
    if m: raw = m.group(1)
    raw = raw.split('?')[0].split('#')[0].strip().strip('/')
    if not re.fullmatch(r'[A-Za-z0-9_-]{20,}', raw):
        raise RuntimeError('Google spreadsheet ID is missing/invalid. Choose a table in /google on Render #1.')
    return raw

def _tab_title(title):
    # R7: deterministic title. Same period/chat updates the same tab instead of creating clutter.
    base = re.sub(r'[\\/?*\[\]:]', ' ', str(title or 'Статьи'))
    base = re.sub(r'\s+', ' ', base).strip(" ' ") or 'Статьи'
    return base[:100].rstrip()

def _cell_value(v):
    if isinstance(v,dict) and v.get('formula'): return {'formulaValue':'='+str(v.get('formula') or '').lstrip('=')}
    if isinstance(v,bool): return {'boolValue':v}
    if isinstance(v,(int,float)) and not isinstance(v,bool): return {'numberValue':float(v)}
    return {'stringValue':str(v or '')}

def _cat_fill(idx):
    # Original vys-262 Google palette/indexing. Columns before the category
    # block use the same pale green-gray header fill as the monolith.
    palette=[(0.78,0.94,0.81),(0.87,0.92,0.97),(0.99,0.89,0.84),(0.89,0.87,0.93),(1.0,0.95,0.8),(0.85,0.92,0.83),(0.81,0.89,0.95),(0.96,0.8,0.8),(0.82,0.88,0.89),(0.92,0.82,0.86),(0.85,0.82,0.91)]
    idx=int(idx or 0)
    if idx >= 3:
        rgb=palette[(idx-3)%len(palette)]
        return {'red':rgb[0],'green':rgb[1],'blue':rgb[2]}
    return {'red':0.92,'green':0.95,'blue':0.9}


def _v262_google_cell_format(row, r_idx, c_idx, max_cols, layout):
    """Google cell format from the final vys-262 monolith."""
    value = row[c_idx - 1] if c_idx - 1 < len(row) else ''
    row_is_blank = not any(str(v if v is not None else '').strip() for v in row)
    first = str(row[0] if row else '').strip().casefold()
    second = str(row[1] if len(row) > 1 else '').strip().casefold()
    fmt = {
        'verticalAlignment': 'TOP',
        'wrapStrategy': 'CLIP' if c_idx == 2 else 'WRAP',
        'borders': {side: {'style': 'SOLID', 'color': {'red': 0.65, 'green': 0.65, 'blue': 0.65}}
                    for side in ('top', 'bottom', 'left', 'right')},
    }
    if isinstance(value, (int, float)) and not isinstance(value, bool) or (isinstance(value, dict) and value.get('formula')):
        fmt['numberFormat'] = {'type': 'NUMBER', 'pattern': '#,##0'}
    if first == 'ars':
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.78, 'green': 0.94, 'blue': 0.81}})
    elif first == 'usd':
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.72, 'green': 0.86, 'blue': 1.0}})
    elif first in {'дата', 'date'}:
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': _cat_fill(c_idx - 1)})
    elif row_is_blank and layout in {'category', 'category_compact'}:
        fmt['backgroundColor'] = {'red': 1.0, 'green': 0.6, 'blue': 0.0}
    elif first in {'расход', 'сумма по статьям'} or second in {'расход', 'сумма по статьям'}:
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': {'red': 1.0, 'green': 0.55, 'blue': 0.55}})
    elif first in {'приход', 'приход за период'} or second in {'приход', 'приход за период'}:
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.55, 'green': 0.78, 'blue': 1.0}})
    elif first in {'остаток с прошлого раза', 'остаток на руках', 'на руках:', 'гомонковые', 'остаток в обороте'} or second in {'остаток с прошлого раза', 'остаток на руках', 'на руках:', 'гомонковые', 'остаток в обороте'}:
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.55, 'green': 0.85, 'blue': 0.55}})
    elif first == 'расход еды на человека в сутки' or second == 'расход еды на человека в сутки':
        fmt.update({'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.74, 'green': 0.82, 'blue': 1.0}})
    elif layout == 'compact' and c_idx in {2, 3} and value not in ('', None):
        fmt['backgroundColor'] = _cat_fill(3 if c_idx == 3 else 2)
    elif layout == 'category_compact' and c_idx >= 3 and value not in ('', None):
        fmt['backgroundColor'] = _cat_fill(c_idx)
    elif layout == 'category' and c_idx >= 4 and value not in ('', None):
        # Exact final vys-262 category indexing.
        fmt['backgroundColor'] = _cat_fill(c_idx - 1)
    return fmt

def create_google_sheet(body):
    """Create or refresh a named tab in the selected owner spreadsheet.

    r2 always tried addSheet, so the next automatic Thu-Wed refresh failed with
    "already exists". r3 reuses an existing tab with the same title and replaces
    only that tab's managed grid. All Google network work stays on Render #2.
    """
    rows = body.get('rows') or []
    layout = str(body.get('layout') or 'category').lower()
    notes_raw = body.get('annotations') or {}
    notes = {}
    for key, val in notes_raw.items():
        try:
            r, c = key.split(',', 1); notes[(int(r), int(c))] = str(val)
        except Exception:
            pass
    token = _google_token()
    spreadsheet_id = _sheet_id(body.get('spreadsheet_id'))
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    info = _google_info()
    meta = requests.get(
        f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}',
        headers=headers,
        params={'fields': 'spreadsheetId,properties.title,sheets.properties(sheetId,title,gridProperties)'},
        timeout=45,
    )
    if meta.status_code >= 300:
        if meta.status_code in (401, 403):
            raise RuntimeError(f'Google Sheets target access {meta.status_code}; share table with {info.get("client_email")}: {meta.text[:400]}')
        raise RuntimeError(f'Google Sheets metadata {meta.status_code}: {meta.text[:400]}')
    payload = meta.json() or {}
    max_cols = max((len(r) for r in rows), default=1)
    row_count = max(100, len(rows) + 20)
    col_count = max(26, max_cols + 3)
    title = _tab_title(body.get('title'))
    existing = None
    for sh in payload.get('sheets') or []:
        props = (sh or {}).get('properties') or {}
        if str(props.get('title') or '') == title:
            existing = props
            break
    if existing:
        sheet_id = int(existing.get('sheetId'))
        grid = existing.get('gridProperties') or {}
        clear_rows = max(row_count, int(grid.get('rowCount') or 0), len(rows) + 5)
        clear_cols = max(col_count, int(grid.get('columnCount') or 0), max_cols + 2)
        # Empty updateCells over a range clears the listed fields. This prevents
        # stale values/notes from a longer previous export from surviving.
        clear_req = {'updateCells': {'range': {
            'sheetId': sheet_id, 'startRowIndex': 0, 'endRowIndex': clear_rows,
            'startColumnIndex': 0, 'endColumnIndex': clear_cols,
        }, 'fields': 'userEnteredValue,note,userEnteredFormat'}}
    else:
        add = requests.post(
            f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}:batchUpdate',
            headers=headers,
            json={'requests': [{'addSheet': {'properties': {'title': title, 'gridProperties': {
                'rowCount': row_count, 'columnCount': col_count,
                'frozenRowCount': 1 if layout in {'compact', 'category_compact'} else 2,
            }}}}]},
            timeout=60,
        )
        if add.status_code >= 300:
            raise RuntimeError(f'Google Sheets add tab {add.status_code}: {add.text[:500]}')
        try:
            sheet_id = int(add.json()['replies'][0]['addSheet']['properties']['sheetId'])
        except Exception as exc:
            raise RuntimeError(f'Google Sheets missing sheetId: {exc}')
        clear_req = None

    cell_rows = []
    for r_idx, row in enumerate(rows, start=1):
        values = []
        for c_idx in range(1, max_cols + 1):
            value = row[c_idx - 1] if c_idx - 1 < len(row) else ''
            cell = {'userEnteredValue': _cell_value(value)}
            note = str(notes.get((r_idx, c_idx)) or '').strip()
            if note:
                cell['note'] = note
            # R9: use the full original vys-262 formatting, not the reduced
            # worker-only coloring introduced by the split.
            cell['userEnteredFormat'] = _v262_google_cell_format(row, r_idx, c_idx, max_cols, layout)
            values.append(cell)
        cell_rows.append({'values': values})

    reqs = []
    if clear_req:
        reqs.append(clear_req)
    reqs.extend([
        {'updateCells': {'range': {'sheetId': sheet_id, 'startRowIndex': 0, 'startColumnIndex': 0},
                         'rows': cell_rows, 'fields': 'userEnteredValue,note,userEnteredFormat'}},
        {'autoResizeDimensions': {'dimensions': {'sheetId': sheet_id, 'dimension': 'COLUMNS',
                                                 'startIndex': 0, 'endIndex': max_cols}}},
    ])
    upd = requests.post(
        f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}:batchUpdate',
        headers=headers, json={'requests': reqs}, timeout=90,
    )
    if upd.status_code >= 300:
        raise RuntimeError(f'Google Sheets update {upd.status_code}: {upd.text[:500]}')

    expected = {(r, c): n.strip() for (r, c), n in notes.items() if n.strip()}
    if expected:
        escaped = title.replace("'", "''")
        verify = requests.get(
            f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}', headers=headers,
            params={'includeGridData': 'true', 'ranges': f"'{escaped}'!A1:ZZ{max(1, len(rows))}",
                    'fields': 'sheets(data(rowData(values(note))))'}, timeout=60,
        )
        if verify.status_code >= 300:
            raise RuntimeError(f'Google Sheets note verify {verify.status_code}: {verify.text[:500]}')
        actual = {}
        try:
            row_data = ((verify.json().get('sheets') or [{}])[0].get('data') or [{}])[0].get('rowData') or []
            for r0, row_obj in enumerate(row_data, start=1):
                for c0, cell in enumerate(row_obj.get('values') or [], start=1):
                    note = str(cell.get('note') or '').strip()
                    if note:
                        actual[(r0, c0)] = note
        except Exception as exc:
            raise RuntimeError(f'Google Sheets note verify parse: {exc}')
        missing = [f'{r}:{c}' for (r, c), note in expected.items() if actual.get((r, c)) != note]
        if missing:
            raise RuntimeError(f'Google Sheets notes not confirmed: {missing[:12]}')
    return f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={sheet_id}'


@app.route('/', methods=['GET','HEAD'])
@app.route('/healthz', methods=['GET','HEAD'])
@app.route('/peer/health', methods=['GET','HEAD'])
def health():
    if request.method == 'HEAD': return '',200
    with STATE_LOCK: state=dict(STATE)
    return {'ok':True,'role':'worker','version':VERSION,'front_configured':bool(front_base()),'mega_configured':bool(os.getenv('MEGA_SESSION') or (os.getenv('MEGA_EMAIL') and os.getenv('MEGA_PASSWORD'))),'google_configured':bool(os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')),'queue_size':JOB_Q.qsize(),'google_queue_size':GOOGLE_Q.qsize(),'state':state},200

@app.route('/internal/event/receipt', methods=['POST'])
def internal_event_receipt_v268():
    if not authorized(): return {'ok':False},404
    wire=request.get_data(cache=False,as_text=False) or b''
    if not wire or len(wire)>env_int('WORKER_EVENT_MAX_WIRE_KB',512,16,4096)*1024: return {'ok':False,'error':'event wire invalid'},413
    try:
        raw=gzip.decompress(wire) if str(request.headers.get('Content-Encoding') or '').lower()=='gzip' else wire
        row=json.loads(raw.decode('utf-8'))
    except Exception as exc: return {'ok':False,'error':f'event decode: {type(exc).__name__}: {str(exc)[:180]}'},400
    if not isinstance(row,dict) or int(row.get('schema') or 0)!=1 or not str(row.get('event_id') or ''): return {'ok':False,'error':'event schema invalid'},400
    eid=str(row.get('event_id'))
    row['state']='received'; row['received_at']=float(row.get('received_at') or time.time()); row['updated_at']=time.time()
    if not _event_local_upsert_v268(row):
        return {'ok':False,'error':'worker local event journal failed'},503
    queued=_event_redis_enqueue_v270(row)
    if not queued:
        rok,rdetail=_event_redis_store_v268(row)
        if not rok:
            with STATE_LOCK: STATE['event_last_error']=str(rdetail)[:220]
            return {'ok':False,'error':'event witness queue+Redis failed: '+str(rdetail)[:180]},503
    with STATE_LOCK:
        STATE['event_received']=int(STATE.get('event_received') or 0)+1; STATE['event_last_at']=time.time(); STATE['event_last_error']=''
    return {'ok':True,'event_id':eid,'state':'received','durable':'worker_local+redis_async'},200

@app.route('/internal/event/commit', methods=['POST'])
def internal_event_commit_v268():
    if not authorized(): return {'ok':False},404
    row=request.get_json(silent=True) or {}; eid=str(row.get('event_id') or row.get('update_id') or '')
    if not eid: return {'ok':False,'error':'event id empty'},400
    state='committed' if str(row.get('state') or '')=='committed' else 'failed_retry'
    row.update({'event_id':eid,'update_id':str(row.get('update_id') or eid),'state':state,'updated_at':time.time()})
    if state=='committed' and not float(row.get('committed_at') or 0.0): row['committed_at']=time.time()
    _event_local_upsert_v268(row); rok,rdetail=_event_redis_store_v268(row)
    if not rok: return {'ok':False,'error':'Redis event update failed: '+str(rdetail)[:180]},503
    with STATE_LOCK:
        if state=='committed': STATE['event_committed']=int(STATE.get('event_committed') or 0)+1
        STATE['event_last_at']=time.time(); STATE['event_last_error']=''
    return {'ok':True,'event_id':eid,'state':state},200

@app.route('/internal/events/pending', methods=['GET'])
def internal_events_pending_v268():
    if not authorized(): return {'ok':False},404
    try: limit=max(1,min(250,int(request.args.get('limit','100') or '100')))
    except Exception: limit=100
    rows=_event_pending_rows_v268(limit)
    return {'ok':True,'events':rows,'count':len(rows)},200

@app.route('/internal/delta', methods=['POST'])
def internal_delta_v267():
    """Apply a compact page delta to the Worker's exact SQLite cache.

    Normal Front updates use this endpoint. A 409 means the Front and Worker bases
    diverged and the Front should perform one full /internal/snapshot/upload rebase.
    """
    if not authorized():
        return {'ok':False},404
    max_wire=env_int('WORKER_DELTA_MAX_WIRE_KB',2048,32,16384)*1024
    wire=request.get_data(cache=False,as_text=False) or b''
    if not wire:
        return {'ok':False,'error':'empty delta'},400
    if len(wire)>max_wire:
        return {'ok':False,'error':f'delta too large: {len(wire)} > {max_wire}'},413
    try:
        raw=gzip.decompress(wire) if str(request.headers.get('Content-Encoding') or '').lower()=='gzip' else wire
        if len(raw)>env_int('WORKER_DELTA_MAX_JSON_MB',16,1,64)*1024*1024:
            return {'ok':False,'error':'delta JSON too large'},413
        payload=json.loads(raw.decode('utf-8'))
    except Exception as exc:
        return {'ok':False,'error':f'delta decode: {type(exc).__name__}: {str(exc)[:240]}'},400
    with DELTA_APPLY_LOCK:
        ok,detail,status=_apply_delta_payload_v267(payload,journal_wire=wire,replay=False)
    if ok:
        with STATE_LOCK:
            return {'ok':True,'status':status,'detail':detail,'state_token':STATE.get('last_state_token'),'state_sha256':STATE.get('last_state_sha256'),'delta_since_checkpoint':STATE.get('delta_since_checkpoint')},200
    with STATE_LOCK:
        STATE['delta_last_error']=str(detail)[:240]
    code=409 if status=='need_full' else 400
    return {'ok':False,'status':status,'error':detail,'worker_sha256':str(STATE.get('last_state_sha256') or '')},code

@app.route('/internal/job', methods=['POST'])
def internal_job():
    global SYNC_PENDING, SYNC_DIRTY, SYNC_DIRTY_REASON
    if not authorized(): return {'ok':False},404
    body=request.get_json(silent=True) or {}; kind=str(body.get('type') or '').strip()
    if kind != 'sync_state': return {'ok':False,'error':'unsupported job type','supported':['sync_state']},400
    token = str(body.get('state_token') or '')[:120]
    with STATE_LOCK:
        last_token = str(STATE.get('last_state_token') or '')
        active_token = str(STATE.get('active_state_token') or '')
        dirty_token = str(STATE.get('dirty_state_token') or '')
        last_full_started = float(STATE.get('last_full_fetch_started') or 0.0)
    try:
        min_full_interval = max(10, min(300, int(os.getenv('WORKER_FULL_REBASE_MIN_INTERVAL_SEC','45') or '45')))
    except Exception:
        min_full_interval = 45
    reason_l = str(body.get('reason') or '').lower()
    force_full = any(x in reason_l for x in ('boot_ready_exact_rebase','shutdown','manual_force'))
    if (not force_full) and last_full_started > 0 and time.time() - last_full_started < min_full_interval:
        with STATE_LOCK:
            STATE['full_fetch_suppressed'] = int(STATE.get('full_fetch_suppressed') or 0) + 1
        return {'ok':True,'status':'full_rebase_rate_limited','retry_after':max(1,int(min_full_interval-(time.time()-last_full_started))),'state_token':token,'queue_size':JOB_Q.qsize()},202
    with SYNC_PENDING_LOCK:
        if not SYNC_PENDING and token and token == last_token:
            with STATE_LOCK: STATE['deduped_sync_requests'] += 1
            return {'ok':True,'status':'up_to_date','state_token':token,'queue_size':JOB_Q.qsize()},200
        if SYNC_PENDING:
            if token and token in {active_token, dirty_token, last_token}:
                with STATE_LOCK: STATE['deduped_sync_requests'] += 1
                return {'ok':True,'status':'coalesced_same','state_token':token,'queue_size':JOB_Q.qsize()},202
            SYNC_DIRTY = True
            SYNC_DIRTY_REASON = str(body.get('reason') or 'coalesced_changes')[:180]
            with STATE_LOCK: STATE['dirty_state_token'] = token
            return {'ok':True,'status':'coalesced_dirty','state_token':token,'queue_size':JOB_Q.qsize()},202
        SYNC_PENDING=True
        with STATE_LOCK: STATE['active_state_token'] = token
    job={'id':secrets.token_hex(8),'type':kind,'reason':str(body.get('reason') or '')[:180],'state_token':token,'created_at':time.time()}
    try: JOB_Q.put_nowait(job)
    except queue.Full:
        with SYNC_PENDING_LOCK: SYNC_PENDING=False
        return {'ok':False,'error':'worker queue full'},503
    return {'ok':True,'status':'queued','job_id':job['id'],'queue_size':JOB_Q.qsize()},202

def _google_status_put(job_id, **values):
    now = time.time()
    with GOOGLE_JOB_LOCK:
        stale = [k for k, row in GOOGLE_JOB_STATUS.items() if now - float((row or {}).get('updated_at') or now) > 86400]
        for key in stale:
            GOOGLE_JOB_STATUS.pop(key, None)
        row = dict(GOOGLE_JOB_STATUS.get(job_id) or {})
        row.update(values)
        row['updated_at'] = now
        GOOGLE_JOB_STATUS[job_id] = row
        return dict(row)


def _notify_front_google_result(job, ok, url='', error=''):
    base, secret = front_base(), peer_secret()
    if not base or not secret:
        return False
    payload = {
        'job_id': str(job.get('id') or ''),
        'ok': bool(ok),
        'url': str(url or ''),
        'error': str(error or '')[:900],
        'title': str((job.get('payload') or {}).get('title') or 'Google Excel')[:220],
        'recipient_chat_id': (job.get('payload') or {}).get('recipient_chat_id'),
        'target_chat_id': (job.get('payload') or {}).get('target_chat_id'),
        'tenant_id': (job.get('payload') or {}).get('tenant_id'),
        'notify_result': bool((job.get('payload') or {}).get('notify_result', True)),
    }
    for attempt in range(1, 4):
        try:
            r = requests.post(base + '/internal/split/google-result', json=payload, headers={'X-Peer-Secret':secret,'User-Agent':'vys-262-worker-google-result'}, timeout=15)
            if 200 <= r.status_code < 300:
                return True
        except Exception:
            pass
        if attempt < 3:
            time.sleep(float(attempt))
    return False


def process_google_job(job):
    jid = str(job.get('id') or '')
    body = dict(job.get('payload') or {})
    _google_status_put(jid, status='running')
    try:
        url = create_google_sheet(body)
        with STATE_LOCK:
            STATE['google_jobs'] += 1
            STATE['google_last_ok'] = time.time()
            STATE['google_last_error'] = ''
        _google_status_put(jid, status='done', ok=True, url=url)
        delivered = _notify_front_google_result(job, True, url=url)
        _google_status_put(jid, callback_delivered=bool(delivered))
        print(f'[GOOGLE JOB] {jid} ok=True callback={delivered}', flush=True)
    except Exception as exc:
        detail = f'{type(exc).__name__}: {str(exc)[:600]}'
        with STATE_LOCK:
            STATE['google_failures'] += 1
            STATE['google_last_error'] = detail[:240]
        _google_status_put(jid, status='done', ok=False, error=detail)
        delivered = _notify_front_google_result(job, False, error=detail)
        _google_status_put(jid, callback_delivered=bool(delivered))
        print(f'[GOOGLE JOB] {jid} ok=False callback={delivered} {detail}', flush=True)


def google_loop():
    while True:
        job = GOOGLE_Q.get()
        try:
            process_google_job(job)
        except Exception as exc:
            print(f'[GOOGLE LOOP ERROR] {type(exc).__name__}: {str(exc)[:300]}', flush=True)
        finally:
            GOOGLE_Q.task_done()


@app.route('/internal/google/sheet', methods=['POST'])
def internal_google_sheet():
    if not authorized():
        return {'ok':False},404
    body = request.get_json(silent=True) or {}
    try:
        spreadsheet_id = _sheet_id(body.get('spreadsheet_id'))
    except Exception as exc:
        return {'ok':False,'error':str(exc)[:600]},400
    body['spreadsheet_id'] = spreadsheet_id
    try:
        recipient_chat_id = int(body.get('recipient_chat_id') or 0)
    except Exception:
        recipient_chat_id = 0
    if not recipient_chat_id:
        return {'ok':False,'error':'recipient_chat_id required'},400
    job_id = str(body.get('job_id') or secrets.token_hex(12)).strip()[:80]
    with GOOGLE_JOB_LOCK:
        existing = dict(GOOGLE_JOB_STATUS.get(job_id) or {})
    if existing:
        status = str(existing.get('status') or 'queued')
        if status == 'done' and existing.get('ok') and existing.get('url'):
            return {'ok':True,'status':'done','job_id':job_id,'url':existing.get('url')},200
        if status == 'done' and not existing.get('ok'):
            return {'ok':False,'status':'done','job_id':job_id,'error':existing.get('error') or 'Google job failed'},409
        return {'ok':True,'status':status,'job_id':job_id,'queue_size':GOOGLE_Q.qsize()},202
    job = {'id':job_id,'type':'google_sheet','created_at':time.time(),'payload':body}
    _google_status_put(job_id, status='queued', ok=None)
    try:
        GOOGLE_Q.put_nowait(job)
    except queue.Full:
        with GOOGLE_JOB_LOCK:
            GOOGLE_JOB_STATUS.pop(job_id, None)
        return {'ok':False,'error':'Google worker queue full'},503
    return {'ok':True,'status':'queued','job_id':job_id,'queue_size':GOOGLE_Q.qsize()},202


@app.route('/internal/google/test', methods=['POST'])
def internal_google_test():
    if not authorized():
        return {'ok':False},404
    body = request.get_json(silent=True) or {}
    try:
        spreadsheet_id = _sheet_id(body.get('spreadsheet_id'))
        token = _google_token()
        info = _google_info()
        headers = {'Authorization':f'Bearer {token}'}
        meta = requests.get(f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}', headers=headers, params={'fields':'spreadsheetId,properties.title'}, timeout=30)
        if meta.status_code >= 300:
            if meta.status_code in (401,403):
                raise RuntimeError(f'Нет доступа к таблице. Расшарьте её {info.get("client_email")} как Редактору. Google: {meta.text[:300]}')
            raise RuntimeError(f'Google Sheets HTTP {meta.status_code}: {meta.text[:350]}')
        title = str((meta.json().get('properties') or {}).get('title') or '')
        return {'ok':True,'spreadsheet_id':spreadsheet_id,'title':title,'service_email':info.get('client_email')},200
    except Exception as exc:
        return {'ok':False,'error':str(exc)[:700]},502

@app.route('/internal/snapshot/upload', methods=['POST'])
def internal_snapshot_upload():
    """Accept the final front SQLite snapshot directly during graceful deploy shutdown.

    Cache is replaced immediately after validation, so a new front instance can restore
    the exact last user state even while MEGA promotion continues in the worker queue.
    """
    if not authorized():
        return {'ok':False},404
    max_bytes = env_int('WORKER_SNAPSHOT_UPLOAD_MAX_MB',64,4,512) * 1024 * 1024
    raw = request.get_data(cache=False, as_text=False) or b''
    if not raw:
        return {'ok':False,'error':'empty snapshot'},400
    if len(raw) > max_bytes:
        return {'ok':False,'error':f'snapshot too large: {len(raw)} > {max_bytes}'},413
    DELTA_APPLY_LOCK.acquire()
    fd, name = tempfile.mkstemp(prefix='vys262_uploaded_', suffix='.sqlite3.gz'); os.close(fd)
    incoming = Path(name)
    try:
        incoming.write_bytes(raw)
        ok, detail, meta = quick_check_gzip(incoming)
        if not ok:
            incoming.unlink(missing_ok=True)
            return {'ok':False,'error':'invalid SQLite snapshot: '+str(detail)[:400]},400
        incoming_revision = float(meta.get('revision') or 0.0)
        with STATE_LOCK:
            current_revision = float(STATE.get('cache_revision') or 0.0)
        cache_exists = CACHE_LATEST.exists()
        if cache_exists and current_revision > 0.0 and (incoming_revision <= 0.0 or incoming_revision < current_revision):
            incoming.unlink(missing_ok=True)
            return {'ok':True,'cached':False,'status':'stale_ignored','revision':incoming_revision,'cache_revision':current_revision},202
        tmp_cache = CACHE_DIR / f'.latest_{secrets.token_hex(6)}.tmp'
        shutil.copy2(incoming, tmp_cache)
        os.replace(tmp_cache, CACHE_LATEST)
        with STATE_LOCK:
            STATE['cache_revision'] = max(current_revision, incoming_revision)
            STATE['last_snapshot_sha256'] = str(meta.get('sha256_gz') or '')
            STATE['last_snapshot_size'] = int(meta.get('size') or len(raw))
            STATE['last_snapshot_at'] = time.time()
            STATE['last_state_token'] = str(request.headers.get('X-Split-State-Token') or STATE.get('last_state_token') or '')[:120]
        try:
            CACHE_DB.unlink(missing_ok=True)
            _ensure_cache_db_v267()
            with STATE_LOCK:
                STATE['last_state_sha256'] = _sha256_file_v267(CACHE_DB) if CACHE_DB.exists() else ''
                STATE['delta_since_checkpoint'] = 0
            redis_store_snapshot(CACHE_LATEST, meta, clear_deltas=True)
        except Exception:
            pass
        job = {
            'id': secrets.token_hex(8), 'type': 'promote_uploaded',
            'reason': str(request.headers.get('X-Snapshot-Reason') or 'front_direct_upload')[:180],
            'created_at': time.time(), 'path': str(incoming),
        }
        try:
            JOB_Q.put_nowait(job)
        except queue.Full:
            # Cache is already exact; keep the file for a short best-effort promoter thread.
            def _late_promote(path=str(incoming)):
                p = Path(path)
                try:
                    mega_promote_snapshot(p)
                finally:
                    p.unlink(missing_ok=True)
            threading.Thread(target=_late_promote, name='vys262-upload-promote', daemon=True).start()
        return {'ok':True,'cached':True,'queued':True,'size':len(raw),'sha256':str(meta.get('sha256_gz') or ''),'revision':float(meta.get('revision') or 0.0)},202
    except Exception as exc:
        incoming.unlink(missing_ok=True)
        return {'ok':False,'error':f'{type(exc).__name__}: {str(exc)[:400]}'},500
    finally:
        try: DELTA_APPLY_LOCK.release()
        except Exception: pass



def _capsule_mega_dir_r20():
    return mega_root().rstrip('/') + '/config_r20'

def _capsule_mega_filename_r20(obj):
    seq,gen,_=_capsule_tuple_r20(obj) if '_capsule_tuple_r20' in globals() else (0,0,0.0)
    stamp=datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S_%f')
    return f'capsule_g{int(gen):010d}_s{int(seq):010d}_{stamp}.json.gz'

def _capsule_mega_rows_r20(limit=30):
    path=_capsule_mega_dir_r20()
    try:
        res=run_cmd(['mega-find',path,'--pattern=capsule_*.json.gz','--type=f'],timeout=60)
        if res.returncode!=0: return []
        rows=[x.strip() for x in (res.stdout or '').splitlines() if x.strip().endswith('.json.gz')]
        return sorted(set(rows),reverse=True)[:max(1,int(limit))]
    except Exception:
        return []

def _capsule_mega_upload_latest_r20():
    if not env_bool('WORKER_CAPSULE_MEGA_ENABLED',True): return True,'MEGA capsule disabled'
    try:
        with CAPSULE_LOCK:
            if not CAPSULE_LOCAL.is_file(): return False,'local capsule missing'
            packed=CAPSULE_LOCAL.read_bytes()
        obj,err=_capsule_decode_r20(packed)
        if not obj: return False,err or 'invalid local capsule'
        with MEGA_LOCK:
            ok,detail=prepare_mega_layout()
            if not ok: return False,detail
            cdir=_capsule_mega_dir_r20()
            if not ensure_mega_dir(cdir): return False,'cannot create capsule MEGA dir'
            work=Path(tempfile.mkdtemp(prefix='vys262_capsule_mega_'))
            try:
                local=work/_capsule_mega_filename_r20(obj); local.write_bytes(packed)
                put=run_cmd(['mega-put',str(local),cdir],timeout=env_int('MEGA_TIMEOUT',180,30,900))
                if put.returncode!=0: return False,'mega-put capsule failed: '+(put.stderr or put.stdout or '')[:180]
                keep=env_int('WORKER_CAPSULE_MEGA_KEEP',10,3,50)
                rows=_capsule_mega_rows_r20(keep+30)
                for old in rows[keep:]:
                    try: run_cmd(['mega-rm',old],timeout=45)
                    except Exception: pass
            finally:
                shutil.rmtree(work,ignore_errors=True)
        CAPSULE_MEGA_STATE['last_ok']=time.time(); CAPSULE_MEGA_STATE['last_error']=''; CAPSULE_MEGA_STATE['uploads']=int(CAPSULE_MEGA_STATE.get('uploads') or 0)+1
        return True,f'MEGA capsule stored gen={_capsule_tuple_r20(obj)[1]} seq={_capsule_tuple_r20(obj)[0]}'
    except Exception as exc:
        detail=f'{type(exc).__name__}: {str(exc)[:200]}'; CAPSULE_MEGA_STATE['last_error']=detail; return False,detail

def _capsule_mega_load_r20():
    if not env_bool('WORKER_CAPSULE_MEGA_ENABLED',True): return {},'MEGA capsule disabled'
    try:
        with MEGA_LOCK:
            ok,detail=prepare_mega_layout()
            if not ok: return {},detail
            cdir=_capsule_mega_dir_r20()
            if not ensure_mega_dir(cdir): return {},'capsule MEGA dir unavailable'
            rows=_capsule_mega_rows_r20(20)
            if not rows: return {},'no MEGA capsule'
            # Filenames are zero-padded by generation then user seq, so reverse sort
            # gives the strongest recent configuration first. Validate more than one
            # file so a partial/corrupt upload cannot break recovery.
            work=Path(tempfile.mkdtemp(prefix='vys262_capsule_get_'))
            try:
                candidates=[]
                for remote in rows[:5]:
                    dl=work/secrets.token_hex(3); dl.mkdir(parents=True,exist_ok=True)
                    get=run_cmd(['mega-get',remote,str(dl)],timeout=env_int('MEGA_TIMEOUT',180,30,900))
                    if get.returncode!=0: continue
                    files=list(dl.rglob('*.json.gz'))+list(dl.rglob('*.gz'))
                    for f in files:
                        obj,err=_capsule_decode_r20(f.read_bytes())
                        if obj: candidates.append(obj)
                obj=_capsule_merge_r20(*candidates) if '_capsule_merge_r20' in globals() else (candidates[0] if candidates else {})
                if obj:
                    CAPSULE_MEGA_STATE['last_ok']=time.time(); CAPSULE_MEGA_STATE['last_error']=''; CAPSULE_MEGA_STATE['restores']=int(CAPSULE_MEGA_STATE.get('restores') or 0)+1
                    return obj,f'MEGA capsule OK gen={_capsule_tuple_r20(obj)[1]} seq={_capsule_tuple_r20(obj)[0]}'
                return {},'MEGA capsule files invalid'
            finally:
                shutil.rmtree(work,ignore_errors=True)
    except Exception as exc:
        detail=f'{type(exc).__name__}: {str(exc)[:200]}'; CAPSULE_MEGA_STATE['last_error']=detail; return {},detail

def _capsule_mega_enqueue_r20():
    if not env_bool('WORKER_CAPSULE_MEGA_ENABLED',True): return False
    try:
        CAPSULE_MEGA_Q.put_nowait(1); return True
    except queue.Full:
        # A queued job reads CAPSULE_LOCAL only when it runs, therefore it naturally
        # uploads the newest coalesced generation even if several changes arrived.
        return True

def _capsule_mega_loop_r20():
    while True:
        CAPSULE_MEGA_Q.get()
        try:
            time.sleep(0.25)
            # Collapse any duplicate wakeups; the file itself is the latest generation.
            try:
                while True: CAPSULE_MEGA_Q.get_nowait(); CAPSULE_MEGA_Q.task_done()
            except queue.Empty: pass
            ok,detail=_capsule_mega_upload_latest_r20()
            if not ok: print('[R20 CAPSULE MEGA] '+detail,flush=True)
        except Exception as exc:
            CAPSULE_MEGA_STATE['last_error']=f'{type(exc).__name__}: {str(exc)[:180]}'
        finally:
            CAPSULE_MEGA_Q.task_done()

def _capsule_decode_r20(raw: bytes):
    try:
        if raw[:2] == b'\x1f\x8b': raw = gzip.decompress(raw)
        obj=json.loads(raw.decode('utf-8'))
        if not isinstance(obj,dict) or str(obj.get('kind') or '')!='vys262_durable_capsule_r20':
            return {}, 'invalid capsule kind'
        return obj, ''
    except Exception as exc:
        return {}, f'{type(exc).__name__}: {str(exc)[:180]}'

def _capsule_tuple_r20(obj):
    return (int((obj or {}).get('user_state_seq') or ((obj or {}).get('user_state') or {}).get('seq') or 0),
            int((obj or {}).get('config_generation') or ((obj or {}).get('config_checkpoint') or {}).get('generation') or 0),
            float((obj or {}).get('saved_at') or 0.0))

def _capsule_merge_r20(*rows):
    rows=[x for x in rows if isinstance(x,dict) and x]
    if not rows: return {}
    merged=dict(max(rows,key=lambda x: float(x.get('saved_at') or 0.0)))
    best_us=max(rows,key=lambda x:int((x.get('user_state') or {}).get('seq') or x.get('user_state_seq') or 0))
    best_cp=max(rows,key=lambda x:int((x.get('config_checkpoint') or {}).get('generation') or x.get('config_generation') or 0))
    merged['user_state']=best_us.get('user_state') or {}; merged['user_state_seq']=int((merged['user_state'] or {}).get('seq') or best_us.get('user_state_seq') or 0)
    merged['config_checkpoint']=best_cp.get('config_checkpoint') or {}; merged['config_generation']=int((merged['config_checkpoint'] or {}).get('generation') or best_cp.get('config_generation') or 0)
    try:
        best_rev=max(rows,key=lambda x:float(((x.get('state_revision') or {}).get('saved_at') or 0.0))); merged['state_revision']=best_rev.get('state_revision') or {}
    except Exception: pass
    merged['saved_at']=max(float(x.get('saved_at') or 0.0) for x in rows)
    merged['kind']='vys262_durable_capsule_r20'; merged['schema']=1
    return merged

def _capsule_load_latest_r20(deep_mega=False):
    rows=[]; detail=[]
    client=_redis_client()
    if client is not None:
        try:
            raw=client.get(_REDIS_CAPSULE_KEY)
            if raw:
                obj,err=_capsule_decode_r20(bytes(raw))
                if obj: rows.append(obj); detail.append('redis')
                elif err: detail.append('redis:'+err)
        except Exception as exc: detail.append('redis:'+str(exc)[:100])
    try:
        if CAPSULE_LOCAL.is_file():
            obj,err=_capsule_decode_r20(CAPSULE_LOCAL.read_bytes())
            if obj: rows.append(obj); detail.append('local')
            elif err: detail.append('local:'+err)
    except Exception as exc: detail.append('local:'+str(exc)[:100])
    if deep_mega or not rows:
        obj,why=_capsule_mega_load_r20()
        if obj: rows.append(obj); detail.append('mega')
        elif why: detail.append('mega:'+why[:100])
    return _capsule_merge_r20(*rows), ','.join(detail)

def _capsule_store_r20(obj:dict, packed:bytes):
    with CAPSULE_LOCK:
        old,_=_capsule_load_latest_r20()
        merged=dict(obj or {})
        if isinstance(old,dict) and old:
            old_seq,old_gen,_old_at=_capsule_tuple_r20(old)
            new_seq,new_gen,_new_at=_capsule_tuple_r20(merged)
            if old_seq > new_seq:
                merged['user_state']=old.get('user_state') or {}; merged['user_state_seq']=old_seq
            if old_gen > new_gen:
                merged['config_checkpoint']=old.get('config_checkpoint') or {}; merged['config_generation']=old_gen
            try:
                if float(((old.get('state_revision') or {}).get('saved_at') or 0.0)) > float(((merged.get('state_revision') or {}).get('saved_at') or 0.0)):
                    merged['state_revision']=old.get('state_revision') or {}
            except Exception: pass
            merged['saved_at']=max(float(old.get('saved_at') or 0.0),float(merged.get('saved_at') or 0.0))
        packed=gzip.compress(json.dumps(merged,ensure_ascii=False,separators=(',',':'),default=str).encode('utf-8'),compresslevel=3)
        new_t=_capsule_tuple_r20(merged)
        tmp=CAPSULE_LOCAL.with_suffix('.tmp'); tmp.write_bytes(packed); os.replace(tmp,CAPSULE_LOCAL)
        client=_redis_client()
        if client is not None:
            meta={'user_state_seq':new_t[0],'config_generation':new_t[1],'saved_at':new_t[2],'size':len(packed),'source':'worker-r20'}
            pipe=client.pipeline(transaction=True); pipe.set(_REDIS_CAPSULE_KEY,packed); pipe.set(_REDIS_CAPSULE_KEY+':meta',json.dumps(meta,separators=(',',':'))); pipe.execute()
        with STATE_LOCK:
            STATE['capsule_seq']=new_t[0]; STATE['capsule_generation']=new_t[1]; STATE['capsule_saved_at']=new_t[2]; STATE['capsule_last_error']=''
        _capsule_mega_enqueue_r20()
        return True, f'stored seq={new_t[0]} gen={new_t[1]}'

@app.route('/internal/capsule',methods=['POST'])
def internal_capsule_r20():
    if not authorized(): return {'ok':False},404
    max_bytes=env_int('WORKER_REDIS_CAPSULE_MAX_MB',8,1,32)*1024*1024
    raw=request.get_data(cache=False)
    if not raw or len(raw)>max_bytes: return {'ok':False,'error':'invalid capsule size'},413
    obj,err=_capsule_decode_r20(raw)
    if not obj: return {'ok':False,'error':err},400
    # Canonical storage remains gzip independent of HTTP server decompression behavior.
    packed=gzip.compress(json.dumps(obj,ensure_ascii=False,separators=(',',':'),default=str).encode('utf-8'),compresslevel=3)
    try:
        ok,detail=_capsule_store_r20(obj,packed)
        return {'ok':ok,'detail':detail,'user_state_seq':_capsule_tuple_r20(obj)[0],'config_generation':_capsule_tuple_r20(obj)[1]},200 if ok else 503
    except Exception as exc:
        with STATE_LOCK: STATE['capsule_last_error']=f'{type(exc).__name__}: {str(exc)[:180]}'
        return {'ok':False,'error':STATE['capsule_last_error']},500

@app.route('/internal/capsule/latest',methods=['GET'])
def internal_capsule_latest_r20():
    if not authorized(): return {'ok':False},404
    deep=str(request.args.get('deep') or '').strip().lower() in {'1','true','yes','on'}
    obj,detail=_capsule_load_latest_r20(deep_mega=deep)
    if not obj: return {'ok':False,'error':'capsule not available','detail':detail},404
    # A deep boot restore also heals local/Redis from MEGA without waiting for the next change.
    if deep:
        try:
            packed0=gzip.compress(json.dumps(obj,ensure_ascii=False,separators=(',',':'),default=str).encode('utf-8'),compresslevel=3)
            with CAPSULE_LOCK:
                tmp=CAPSULE_LOCAL.with_suffix('.tmp'); tmp.write_bytes(packed0); os.replace(tmp,CAPSULE_LOCAL)
            client=_redis_client()
            if client is not None:
                pipe=client.pipeline(transaction=True); pipe.set(_REDIS_CAPSULE_KEY,packed0); pipe.set(_REDIS_CAPSULE_KEY+':meta',json.dumps({'user_state_seq':_capsule_tuple_r20(obj)[0],'config_generation':_capsule_tuple_r20(obj)[1],'saved_at':_capsule_tuple_r20(obj)[2],'size':len(packed0),'source':'worker-r20-deep'},separators=(',',':'))); pipe.execute()
        except Exception: pass
    packed=gzip.compress(json.dumps(obj,ensure_ascii=False,separators=(',',':'),default=str).encode('utf-8'),compresslevel=3)
    resp=Response(packed,status=200,mimetype='application/gzip')
    resp.headers['X-Capsule-User-Seq']=str(_capsule_tuple_r20(obj)[0]); resp.headers['X-Capsule-Config-Generation']=str(_capsule_tuple_r20(obj)[1])
    return resp

@app.route('/internal/restore/latest', methods=['GET'])
def internal_restore_latest():
    if not authorized(): return {'ok':False},404
    cache_max_age=env_int('WORKER_RESTORE_CACHE_MAX_AGE_SEC',120,0,3600)
    if not CACHE_LATEST.exists():
        redis_ok, redis_detail = redis_load_snapshot_to_cache()
        if not redis_ok:
            started=_schedule_restore_refresh()
            return {'ok':False,'error':'restore cache not ready','redis':redis_detail,'refresh_queued':bool(started)},503
    # Never serve corrupt cache just because the file exists.
    ok, detail, meta = quick_check_gzip(CACHE_LATEST)
    if not ok:
        CACHE_LATEST.unlink(missing_ok=True)
        redis_ok, redis_detail = redis_load_snapshot_to_cache()
        if not redis_ok:
            started=_schedule_restore_refresh()
            return {'ok':False,'error':'restore cache invalid','detail':detail,'redis':redis_detail,'refresh_queued':bool(started)},503
        ok, detail, meta = quick_check_gzip(CACHE_LATEST)
        if not ok:
            return {'ok':False,'error':'Redis restore cache invalid after reload'},503
    age=max(0.0,time.time()-CACHE_LATEST.stat().st_mtime)
    if cache_max_age <= 0 or age > cache_max_age:
        _schedule_restore_refresh()
    payload=CACHE_LATEST.read_bytes()
    resp=Response(payload,status=200,mimetype='application/gzip')
    resp.headers['Content-Disposition']='attachment; filename="latest_bot_state.sqlite3.gz"'
    resp.headers['X-Worker-Version']=TRANSPORT_VERSION
    resp.headers['X-SHA256']=hashlib.sha256(payload).hexdigest()
    resp.headers['X-Cache-Age-Sec']=str(int(age))
    resp.headers['X-State-Revision']=str(float(meta.get('revision') or 0.0))
    return resp

@app.route('/internal/status', methods=['GET'])
def internal_status():
    if not authorized(): return {'ok':False},404
    with STATE_LOCK: return {'ok':True,'version':VERSION,'queue_size':JOB_Q.qsize(),'google_queue_size':GOOGLE_Q.qsize(),'state':dict(STATE)},200

# ---------------------------------------------------------------------------
# R7 heavy file exports: CSV/XLSX/Drive live here, away from Telegram webhooks.
FILE_Q = queue.Queue(maxsize=12)
FILE_JOB_LOCK = threading.RLock()
FILE_JOB_STATUS = {}
FILE_DIR = CACHE_DIR / 'exports'
FILE_DIR.mkdir(parents=True, exist_ok=True)
try:
    STATE.update({'file_jobs':0, 'file_failures':0, 'file_last_ok':0.0, 'file_last_error':''})
except Exception:
    pass


def _file_status_put(job_id, **values):
    now = time.time()
    with FILE_JOB_LOCK:
        for key, row in list(FILE_JOB_STATUS.items()):
            if now - float((row or {}).get('updated_at') or now) > 7200:
                old = FILE_JOB_STATUS.pop(key, {}) or {}
                try:
                    path = Path(str(old.get('path') or ''))
                    if path.is_file(): path.unlink(missing_ok=True)
                except Exception: pass
        row = dict(FILE_JOB_STATUS.get(job_id) or {})
        row.update(values); row['updated_at'] = now
        FILE_JOB_STATUS[job_id] = row
        return dict(row)


def _safe_export_name(value, ext):
    raw = str(value or f'export.{ext}').replace('\\','_').replace('/','_').strip()
    raw = re.sub(r'[\x00-\x1f<>:"|?*]+','_',raw)[:180] or f'export.{ext}'
    if not raw.lower().endswith('.'+ext): raw += '.'+ext
    return raw


def _xlsx_value(value):
    if isinstance(value, dict):
        formula = str(value.get('formula') or '').strip()
        if formula:
            return '=' + formula.lstrip('=')
        if 'value' in value:
            return value.get('value')
    return value


def _r10_rgb_hex(rgb):
    if not isinstance(rgb, dict): return None
    try:
        vals=[]
        for key in ('red','green','blue'):
            raw=float(rgb.get(key,0.0) or 0.0)
            vals.append(max(0,min(255,round(raw*255))))
        return ''.join(f'{v:02X}' for v in vals)
    except Exception:
        return None

def _r10_apply_v262_xlsx_style(cell, row, r_idx, c_idx, max_cols, layout):
    """R16 canonical palette for every XLSX file and Google Sheets path."""
    try:
        fmt=_v262_google_cell_format(row,r_idx,c_idx,max_cols,layout) or {}
    except Exception:
        fmt={}
    fill=_r10_rgb_hex(fmt.get('backgroundColor'))
    if fill: cell.fill=PatternFill('solid',fgColor=fill)
    tf=fmt.get('textFormat') or {}
    if tf.get('bold'): cell.font=Font(bold=True)
    wrap = str(fmt.get('wrapStrategy') or '').upper() != 'CLIP'
    h=str(fmt.get('horizontalAlignment') or '').lower() or None
    v=str(fmt.get('verticalAlignment') or 'TOP').lower()
    cell.alignment=Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    nf=fmt.get('numberFormat') or {}
    if nf.get('pattern'): cell.number_format=str(nf.get('pattern'))
    borders=fmt.get('borders') or {}
    def side(name):
        spec=borders.get(name) or {}
        if not spec: return Side(style=None)
        color=_r10_rgb_hex(spec.get('color')) or 'A6A6A6'
        style='thin' if str(spec.get('style') or '').upper() != 'NONE' else None
        return Side(style=style,color=color)
    if borders:
        cell.border=Border(left=side('left'),right=side('right'),top=side('top'),bottom=side('bottom'))

def _render_export_file(body, job_id):
    ftype = 'xlsx' if str(body.get('file_type') or '').lower() in {'xlsx','xlsxstat','excel'} else 'csv'
    filename = _safe_export_name(body.get('filename'), ftype)
    path = FILE_DIR / f'{job_id}.{ftype}'
    rows = body.get('rows') or []
    if ftype == 'csv':
        with open(path,'w',newline='',encoding='utf-8-sig') as fh:
            w=csv.writer(fh)
            for row in rows: w.writerow(list(row or []))
        return path, filename
    wb=Workbook(); ws=wb.active; ws.title=str(body.get('sheet_name') or 'Экспорт')[:31] or 'Экспорт'
    annotations={}
    for key,val in (body.get('annotations') or {}).items():
        try:
            rr,cc=str(key).split(',',1); annotations[(int(rr),int(cc))]=str(val)
        except Exception: pass
    raw_layout = body.get('category_layout')
    layout = 'category' if raw_layout is True else (str(raw_layout or body.get('layout') or 'category').lower())
    style = str(body.get('style') or 'old')
    max_cols=max((len(list(row or [])) for row in rows), default=1)
    for r_idx,row in enumerate(rows,start=1):
        vals=list(row or [])
        padded=vals + [''] * max(0,max_cols-len(vals))
        for c_idx,value in enumerate(padded,start=1):
            cell=ws.cell(r_idx,c_idx,value=_xlsx_value(value))
            _r10_apply_v262_xlsx_style(cell,padded,r_idx,c_idx,max_cols,layout)
            note=annotations.get((r_idx,c_idx),'').strip()
            if note and style in {'new_comments','new_notes','old','new_plain'}:
                cell.comment=Comment(note,'Telegram Finance Bot')
    ws.freeze_panes='A2'
    for col in range(1,max_cols+1):
        letter=get_column_letter(col)
        width=10
        for cell in ws[letter][:min(ws.max_row,300)]:
            try: width=max(width,min(48,len(str(cell.value or ''))+2))
            except Exception: pass
        ws.column_dimensions[letter].width=width
    wb.save(path)
    return path, filename


def _drive_upload_file(path: Path, filename: str, folder_id: str=''):
    token=_google_token(); headers={'Authorization':f'Bearer {token}'}
    metadata={'name':filename}
    folder=str(folder_id or '').strip()
    if folder:
        folder=_sheet_id(folder) if '/spreadsheets/' in folder else re.sub(r'^.*/folders/','',folder).split('?')[0].split('#')[0].strip('/')
        if not re.fullmatch(r'[A-Za-z0-9_-]{10,}',folder): raise RuntimeError('Google Drive folder ID invalid')
        metadata['parents']=[folder]
    mime=mimetypes.guess_type(filename)[0] or 'application/octet-stream'
    with open(path,'rb') as fh:
        r=requests.post('https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&fields=id,name,webViewLink', headers=headers,
            files={'metadata':('metadata',json.dumps(metadata),'application/json; charset=UTF-8'),'file':(filename,fh,mime)}, timeout=120)
    if r.status_code >= 300: raise RuntimeError(f'Google Drive upload {r.status_code}: {r.text[:500]}')
    payload=r.json() or {}; fid=str(payload.get('id') or '')
    return str(payload.get('webViewLink') or (f'https://drive.google.com/file/d/{fid}/view' if fid else ''))


def _notify_front_export_result(job, ok, **extra):
    base,secret=front_base(),peer_secret()
    if not base or not secret: return False
    body=dict(job.get('payload') or {})
    payload={'job_id':str(job.get('id') or ''),'ok':bool(ok),'error':str(extra.get('error') or '')[:900],
        'url':str(extra.get('url') or ''),'filename':str(extra.get('filename') or body.get('filename') or ''),
        'file_type':str(body.get('file_type') or ''),'delivery':str(body.get('delivery') or 'chat'),
        'recipient_chat_id':body.get('recipient_chat_id'),'target_chat_id':body.get('target_chat_id'),
        'tenant_id':body.get('tenant_id'),'label':str(body.get('label') or ''),'chat_name':str(body.get('chat_name') or ''),'caption':str(body.get('caption') or '')[:1000]}
    for attempt in range(3):
        try:
            r=requests.post(base+'/internal/split/export-result',json=payload,headers={'X-Peer-Secret':secret,'User-Agent':'vys-262-worker-export-r7'},timeout=15)
            if 200 <= r.status_code < 300: return True
        except Exception: pass
        time.sleep(attempt+1)
    return False


def process_file_job(job):
    jid=str(job.get('id') or ''); body=dict(job.get('payload') or {})
    _file_status_put(jid,status='running')
    try:
        path,filename=_render_export_file(body,jid)
        delivery=str(body.get('delivery') or 'chat')
        url=''
        if delivery == 'drive':
            url=_drive_upload_file(path,filename,str(body.get('drive_folder_id') or ''))
        with STATE_LOCK:
            STATE['file_jobs']=int(STATE.get('file_jobs') or 0)+1; STATE['file_last_ok']=time.time(); STATE['file_last_error']=''
        _file_status_put(jid,status='done',ok=True,path=str(path),filename=filename,url=url,delivery=delivery)
        delivered=_notify_front_export_result(job,True,url=url,filename=filename)
        _file_status_put(jid,callback_delivered=bool(delivered))
        if delivery == 'drive': path.unlink(missing_ok=True)
        print(f'[EXPORT JOB] {jid} ok=True delivery={delivery} callback={delivered}',flush=True)
    except Exception as exc:
        detail=f'{type(exc).__name__}: {str(exc)[:700]}'
        with STATE_LOCK:
            STATE['file_failures']=int(STATE.get('file_failures') or 0)+1; STATE['file_last_error']=detail[:240]
        _file_status_put(jid,status='done',ok=False,error=detail)
        _notify_front_export_result(job,False,error=detail)
        print(f'[EXPORT JOB] {jid} ok=False {detail}',flush=True)


def file_loop():
    while True:
        job=FILE_Q.get()
        try: process_file_job(job)
        except Exception as exc: print(f'[EXPORT LOOP ERROR] {type(exc).__name__}: {str(exc)[:300]}',flush=True)
        finally: FILE_Q.task_done()


@app.route('/internal/google/info',methods=['GET'])
def internal_google_info_r7():
    if not authorized(): return {'ok':False},404
    try:
        info=_google_info()
        return {'ok':True,'service_email':str(info.get('client_email') or ''),'configured':True},200
    except Exception as exc:
        return {'ok':False,'configured':False,'error':str(exc)[:500]},502


@app.route('/internal/export/file',methods=['POST'])
def internal_export_file_r7():
    if not authorized(): return {'ok':False},404
    body=request.get_json(silent=True) or {}
    try: cid=int(body.get('recipient_chat_id') or 0)
    except Exception: cid=0
    if not cid: return {'ok':False,'error':'recipient_chat_id required'},400
    rows=body.get('rows') or []
    if not isinstance(rows,list) or len(rows)>100000: return {'ok':False,'error':'invalid/too many rows'},400
    jid=str(body.get('job_id') or secrets.token_hex(12)).strip()[:80]
    with FILE_JOB_LOCK: existing=dict(FILE_JOB_STATUS.get(jid) or {})
    if existing:
        return {'ok':bool(existing.get('ok',True)),'status':existing.get('status') or 'queued','job_id':jid},200 if existing.get('status')=='done' else 202
    job={'id':jid,'type':'file_export','created_at':time.time(),'payload':body}
    _file_status_put(jid,status='queued',ok=None)
    try: FILE_Q.put_nowait(job)
    except queue.Full:
        with FILE_JOB_LOCK: FILE_JOB_STATUS.pop(jid,None)
        return {'ok':False,'error':'file worker queue full'},503
    return {'ok':True,'status':'queued','job_id':jid,'queue_size':FILE_Q.qsize()},202


@app.route('/internal/export/file/<job_id>',methods=['GET'])
def internal_export_download_r7(job_id):
    if not authorized(): return {'ok':False},404
    jid=str(job_id or '')[:80]
    with FILE_JOB_LOCK: row=dict(FILE_JOB_STATUS.get(jid) or {})
    if not row or row.get('status')!='done' or not row.get('ok'): return {'ok':False,'error':'file not ready'},404
    path=Path(str(row.get('path') or ''))
    if not path.is_file(): return {'ok':False,'error':'file expired'},410
    return send_file(path,as_attachment=True,download_name=str(row.get('filename') or path.name),max_age=0)

threading.Thread(target=file_loop,name='vys262-worker-files-r7',daemon=True).start()
threading.Thread(target=_capsule_mega_loop_r20,name='vys262-worker-capsule-mega-r20',daemon=True).start()

threading.Thread(target=_event_redis_flush_loop_v270,name='vys262-worker-event-redis-r15',daemon=True).start()

threading.Thread(target=worker_loop,name='vys262-worker-jobs',daemon=True).start()
threading.Thread(target=google_loop,name='vys262-worker-google',daemon=True).start()
threading.Thread(target=peer_loop,name='vys262-worker-peer',daemon=True).start()
try:
    _r6_redis_ok, _r6_redis_detail = redis_load_snapshot_to_cache()
    print(f'[R6 RESTORE CACHE] redis ok={_r6_redis_ok} {_r6_redis_detail}', flush=True)
    if _r6_redis_ok:
        _r12_replay_ok, _r12_replay_detail = redis_replay_deltas_v267()
        print(f'[R12 DELTA REPLAY] ok={_r12_replay_ok} {_r12_replay_detail}', flush=True)
except Exception as _r6_exc:
    print(f'[R6 RESTORE CACHE] redis error={type(_r6_exc).__name__}: {str(_r6_exc)[:180]}', flush=True)
threading.Thread(target=_restore_refresh_background,name='vys262-worker-mega-warmup',daemon=True).start()
threading.Thread(target=_checkpoint_loop_v267,name='vys262-worker-checkpoint-r13',daemon=True).start()
threading.Thread(target=_event_reconcile_loop_v268,name='vys262-worker-events-r13',daemon=True).start()
threading.Thread(target=_reconcile_hash_loop_v268,name='vys262-worker-reconcile-r13',daemon=True).start()

if __name__ == '__main__':
    app.run(host='0.0.0.0',port=env_int('PORT',10000,1,65535),threaded=True)
# v262
